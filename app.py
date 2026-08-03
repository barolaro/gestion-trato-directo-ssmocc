from __future__ import annotations

import base64
import csv
import gzip
import hashlib
import hmac
import io
import json
import re
import secrets as pysecrets
import unicodedata
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import gspread
from google.oauth2.service_account import Credentials


st.set_page_config(
    page_title="Inteligencia de Adquisiciones SSMOCC",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

BASE_DIR = Path(__file__).resolve().parent
HTML_FILES = (
    BASE_DIR / "index.html",
    BASE_DIR / "index_dashboard_final_corregido_gestion.html",
)

st.markdown(
    """
    <style>
      #MainMenu, footer, [data-testid="stHeader"], [data-testid="stToolbar"] {
        display: none !important;
      }
      [data-testid="stAppViewContainer"],
      [data-testid="stMain"],
      [data-testid="stMainBlockContainer"],
      .block-container {
        width: 100% !important;
        max-width: 100vw !important;
        padding: 0 !important;
        margin: 0 !important;
      }
      [data-testid="stVerticalBlock"] { gap: 0 !important; }
      iframe[title="streamlit.components.v1.html"] {
        width: 100% !important;
        max-width: 100% !important;
        border: 0 !important;
        display: block !important;
      }
    </style>
    """,
    unsafe_allow_html=True,
)


# -----------------------------------------------------------------------------
# GOOGLE SHEETS · CAPA DE DATOS
# -----------------------------------------------------------------------------
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

SHEET_SCHEMAS: dict[str, list[str]] = {
    "establecimientos": ["id", "nombre", "codigo", "activo"],
    "contratos": [
        "id", "establecimiento_id", "licitacion", "monto_adjudicado",
        "fecha_adjudicacion", "duracion_meses", "anticipacion_renovacion",
        "estado", "responsable", "observaciones", "ultima_actualizacion",
        "estado_revision", "enviado_revision", "actualizado_por",
    ],
    "planes": [
        "id", "nombre_archivo", "reporte", "periodo", "fecha_publicacion",
        "establecimientos", "rojos", "amarillos", "verdes", "url_archivo",
    ],
    "plan_trabajo": [
        "id", "establecimiento_id", "reporte", "periodo",
        "fecha_publicacion", "nivel", "acciones", "responsable",
        "fecha_compromiso", "estado", "observaciones",
    ],
    "datos_dashboard": ["id", "orden", "contenido"],
    "cargas_mensuales": [
        "id", "nombre_archivo", "fecha_carga", "registros",
        "periodo_min", "periodo_max", "establecimientos",
    ],
    "resultados_minsal": [
        "id", "establecimiento_id", "reporte", "periodo", "codigo_deis",
        "denominador", "numerador", "porcentaje_td", "nivel",
        "nombre_archivo", "fecha_carga",
    ],
    "usuarios_establecimientos": [
        "id", "usuario", "clave_hash", "establecimiento_id",
        "nombre", "activo", "creado", "ultima_conexion",
    ],
    "historial_cambios": [
        "id", "contrato_id", "establecimiento_id", "usuario",
        "accion", "detalle", "fecha",
    ],
}

DEFAULT_ESTABLISHMENTS = [
    [1, "Hospital San Juan de Dios", "HSJD", "TRUE"],
    [2, "Instituto Traumatológico", "IT", "TRUE"],
    [3, "Hospital Dr. Félix Bulnes Cerda", "HFBC", "TRUE"],
    [4, "Hospital de Talagante", "HTAL", "TRUE"],
    [5, "Hospital de Peñaflor", "HPE", "TRUE"],
    [6, "Hospital de Melipilla", "HMEL", "TRUE"],
    [7, "Hospital de Curacaví", "HCUR", "TRUE"],
    [8, "CRS Salvador Allende", "CRS", "TRUE"],
    [9, "SSMOCC Dirección", "DSS", "TRUE"],
]


def secret(name: str, default: str = "") -> str:
    try:
        value = st.secrets.get(name, default)
        return str(value).strip() if value is not None else default
    except Exception:
        return default


def _coerce(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        if value.lower() == "true":
            return True
        if value.lower() == "false":
            return False
    return value


def _to_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return value


def _same_value(left: Any, right: Any) -> bool:
    if left is None or right is None:
        return left is right
    return str(left).strip() == str(right).strip()


class SheetResult:
    def __init__(self, data: list[dict[str, Any]]):
        self.data = data


class SheetQuery:
    def __init__(self, worksheet: gspread.Worksheet):
        self._ws = worksheet
        self._operation = "select"
        self._columns = "*"
        self._payload: Any = None
        self._filters: list[tuple[str, str, Any]] = []
        self._limit: int | None = None

    def select(self, columns: str = "*") -> "SheetQuery":
        self._operation, self._columns = "select", columns
        return self

    def insert(self, payload: Any) -> "SheetQuery":
        self._operation, self._payload = "insert", payload
        return self

    def update(self, payload: dict[str, Any]) -> "SheetQuery":
        self._operation, self._payload = "update", payload
        return self

    def delete(self) -> "SheetQuery":
        self._operation = "delete"
        return self

    def eq(self, column: str, value: Any) -> "SheetQuery":
        self._filters.append(("eq", column, value))
        return self

    def neq(self, column: str, value: Any) -> "SheetQuery":
        self._filters.append(("neq", column, value))
        return self

    def limit(self, count: int) -> "SheetQuery":
        self._limit = count
        return self

    def _rows(self) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for index, record in enumerate(self._ws.get_all_records()):
            row = {key: _coerce(value) for key, value in record.items()}
            if row.get("id") not in ("", None):
                try:
                    row["id"] = int(row["id"])
                except (TypeError, ValueError):
                    pass
            row["__rownum__"] = index + 2
            rows.append(row)
        return rows

    def _matches(self, row: dict[str, Any]) -> bool:
        for kind, column, expected in self._filters:
            equal = _same_value(row.get(column), expected)
            if (kind == "eq" and not equal) or (kind == "neq" and equal):
                return False
        return True

    def execute(self) -> SheetResult:
        if self._operation == "select":
            rows = [row for row in self._rows() if self._matches(row)]
            for row in rows:
                row.pop("__rownum__", None)
            if self._limit is not None:
                rows = rows[:self._limit]
            if self._columns != "*":
                columns = [column.strip() for column in self._columns.split(",")]
                rows = [{column: row.get(column) for column in columns} for row in rows]
            return SheetResult(rows)

        headers = self._ws.row_values(1)
        if not headers:
            raise RuntimeError(f"La hoja {self._ws.title} no tiene encabezados.")

        if self._operation == "insert":
            payloads = self._payload if isinstance(self._payload, list) else [self._payload]
            ids = [row["id"] for row in self._rows() if isinstance(row.get("id"), int)]
            next_id = max(ids, default=0) + 1
            values = []
            for payload in payloads:
                record = dict(payload)
                if "id" in headers and not record.get("id"):
                    record["id"] = next_id
                    next_id += 1
                values.append([_to_cell(record.get(header, "")) for header in headers])
            if values:
                self._ws.append_rows(values, value_input_option="RAW")
            return SheetResult(payloads)

        if self._operation == "update":
            targets = [row for row in self._rows() if self._matches(row)]
            for row in targets:
                row_number = row["__rownum__"]
                current = self._ws.row_values(row_number)
                values = [
                    _to_cell(self._payload[header])
                    if header in self._payload
                    else (current[index] if index < len(current) else "")
                    for index, header in enumerate(headers)
                ]
                self._ws.update(
                    range_name=f"A{row_number}",
                    values=[values],
                    value_input_option="RAW",
                )
            return SheetResult([self._payload])

        if self._operation == "delete":
            targets = [row for row in self._rows() if self._matches(row)]
            for row in sorted(targets, key=lambda item: item["__rownum__"], reverse=True):
                self._ws.delete_rows(row["__rownum__"])
            return SheetResult([])

        return SheetResult([])


class SheetClient:
    def __init__(self, spreadsheet: gspread.Spreadsheet):
        self._spreadsheet = spreadsheet
        self._worksheets: dict[str, gspread.Worksheet] = {}

    def table(self, name: str) -> SheetQuery:
        if name not in self._worksheets:
            try:
                worksheet = self._spreadsheet.worksheet(name)
            except gspread.WorksheetNotFound:
                headers = SHEET_SCHEMAS.get(name)
                if not headers:
                    raise RuntimeError(f"No existe la hoja '{name}'.")
                worksheet = self._spreadsheet.add_worksheet(
                    title=name, rows=200, cols=len(headers)
                )
                worksheet.update(range_name="A1", values=[headers], value_input_option="RAW")
                if name == "establecimientos":
                    worksheet.append_rows(
                        DEFAULT_ESTABLISHMENTS, value_input_option="RAW"
                    )
            expected_headers = SHEET_SCHEMAS.get(name, [])
            existing_headers = worksheet.row_values(1)
            missing_headers = [
                header for header in expected_headers if header not in existing_headers
            ]
            if missing_headers:
                first_column = len(existing_headers) + 1
                required_columns = len(existing_headers) + len(missing_headers)
                if worksheet.col_count < required_columns:
                    worksheet.resize(cols=required_columns)
                worksheet.update(
                    range_name=gspread.utils.rowcol_to_a1(1, first_column),
                    values=[missing_headers],
                    value_input_option="RAW",
                )
            self._worksheets[name] = worksheet
        return SheetQuery(self._worksheets[name])

    def replace_records(
        self, name: str, records: list[dict[str, Any]]
    ) -> None:
        """Reemplaza una tabla completa con una sola escritura masiva."""
        if name not in self._worksheets:
            try:
                self._worksheets[name] = self._spreadsheet.worksheet(name)
            except gspread.WorksheetNotFound:
                headers = SHEET_SCHEMAS[name]
                worksheet = self._spreadsheet.add_worksheet(
                    title=name, rows=200, cols=len(headers)
                )
                self._worksheets[name] = worksheet
        worksheet = self._worksheets[name]
        headers = SHEET_SCHEMAS[name]
        values = [
            [_to_cell(record.get(header, "")) for header in headers]
            for record in records
        ]
        required_rows = max(2, len(values) + 1)
        if worksheet.row_count < required_rows or worksheet.col_count < len(headers):
            worksheet.resize(
                rows=max(worksheet.row_count, required_rows),
                cols=max(worksheet.col_count, len(headers)),
            )
        worksheet.clear()
        worksheet.update(
            range_name="A1",
            values=[headers] + values,
            value_input_option="RAW",
        )


@st.cache_resource
def _spreadsheet() -> gspread.Spreadsheet | None:
    account_info: dict[str, Any] = {}
    try:
        account_info = dict(st.secrets["gcp_service_account"])
    except Exception:
        raw_json = secret("GCP_SERVICE_ACCOUNT_JSON")
        if raw_json:
            try:
                account_info = json.loads(raw_json)
            except json.JSONDecodeError as exc:
                raise RuntimeError(
                    "GCP_SERVICE_ACCOUNT_JSON no contiene un JSON válido."
                ) from exc
    sheet_id = secret("GSHEET_ID") or secret("GOOGLE_SHEET_ID")
    if not account_info or not sheet_id:
        return None
    credentials = Credentials.from_service_account_info(
        account_info, scopes=GOOGLE_SCOPES
    )
    return gspread.authorize(credentials).open_by_key(sheet_id)


@st.cache_resource
def _sheet_client() -> SheetClient | None:
    spreadsheet = _spreadsheet()
    return SheetClient(spreadsheet) if spreadsheet is not None else None


def public_client() -> SheetClient:
    client = _sheet_client()
    if client is None:
        raise RuntimeError(
            "Faltan gcp_service_account y GSHEET_ID en Streamlit Secrets."
        )
    return client


def service_client() -> SheetClient | None:
    return _sheet_client()



@st.cache_resource
def ensure_optional_sheet(name: str) -> bool:
    """Crea una hoja nueva sin permitir que su ausencia bloquee el dashboard."""
    spreadsheet = _spreadsheet()
    headers = SHEET_SCHEMAS.get(name, [])
    if spreadsheet is None or not headers:
        return False
    try:
        worksheet = spreadsheet.worksheet(name)
    except gspread.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(
            title=name, rows=200, cols=len(headers)
        )
        worksheet.update(
            range_name="A1", values=[headers], value_input_option="RAW"
        )
    return worksheet is not None

@st.cache_resource
def ensure_all_sheets() -> bool:
    """Crea de una vez las pestañas faltantes con una sola lectura de metadatos."""
    spreadsheet = _spreadsheet()
    if spreadsheet is None:
        return False
    existing = {worksheet.title: worksheet for worksheet in spreadsheet.worksheets()}
    for name, headers in SHEET_SCHEMAS.items():
        if name in existing:
            continue
        worksheet = spreadsheet.add_worksheet(
            title=name, rows=200, cols=len(headers)
        )
        worksheet.update(
            range_name="A1", values=[headers], value_input_option="RAW"
        )
        existing[name] = worksheet
    return True


def safe_read(table: str, columns: str = "*") -> list[dict[str, Any]]:
    try:
        result = public_client().table(table).select(columns).execute()
        return list(result.data or [])
    except Exception as exc:
        st.warning(f"No fue posible leer {table}: {exc}")
        return []


def _records_from_values(values: list[list[Any]]) -> list[dict[str, Any]]:
    if not values:
        return []
    headers = [str(value).strip() for value in values[0]]
    records: list[dict[str, Any]] = []
    for source in values[1:]:
        record = {
            header: _coerce(source[index] if index < len(source) else "")
            for index, header in enumerate(headers)
            if header
        }
        if record.get("id") not in ("", None):
            try:
                record["id"] = int(record["id"])
            except (TypeError, ValueError):
                pass
        records.append(record)
    return records


def batch_read_tables() -> dict[str, list[dict[str, Any]]]:
    """Lee todas las pestañas con una única solicitud batchGet."""
    spreadsheet = _spreadsheet()
    if spreadsheet is None:
        raise RuntimeError("No existe conexión configurada con Google Sheets.")
    table_names = list(SHEET_SCHEMAS)
    ranges = [f"'{name}'!A:ZZ" for name in table_names]
    response = spreadsheet.values_batch_get(
        ranges,
        params={"valueRenderOption": "UNFORMATTED_VALUE"},
    )
    value_ranges = response.get("valueRanges", [])
    tables: dict[str, list[dict[str, Any]]] = {}
    for index, name in enumerate(table_names):
        values = (
            value_ranges[index].get("values", [])
            if index < len(value_ranges) else []
        )
        tables[name] = _records_from_values(values)
    return tables


@st.cache_data(ttl=60, show_spinner=False)
def load_data() -> dict[str, list[dict[str, Any]]]:
    try:
        ensure_all_sheets()
        tables = batch_read_tables()
    except Exception as exc:
        st.warning(
            "Google Sheets alcanzó temporalmente su límite de consultas. "
            "El dashboard continuará mostrando la última base incorporada. "
            f"Detalle: {exc}"
        )
        tables = {name: [] for name in SHEET_SCHEMAS}

    establishments = tables.get("establecimientos", [])
    chunks = tables.get("datos_dashboard", [])
    chunks.sort(key=lambda row: int(row.get("orden") or 0))
    dataset_gzip_b64 = "".join(
        str(row.get("contenido") or "") for row in chunks
    )
    return {
        "establecimientos": [
            row for row in establishments
            if row.get("activo", True) not in (False, "")
        ],
        "contratos": tables.get("contratos", []),
        "planes": tables.get("planes", []),
        "plan_trabajo": tables.get("plan_trabajo", []),
        "cargas_mensuales": tables.get("cargas_mensuales", []),
        "resultados_minsal": tables.get("resultados_minsal", []),
        "usuarios_establecimientos": tables.get(
            "usuarios_establecimientos", []
        ),
        "historial_cambios": tables.get("historial_cambios", []),
        "dataset_gzip_b64": dataset_gzip_b64,
    }

# -----------------------------------------------------------------------------
# NORMALIZACIÓN Y NOMBRES
# -----------------------------------------------------------------------------
def normalize(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = "".join(
        char
        for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def load_dashboard_dataset_b64() -> str:
    """Reconstruye el dataset comprimido guardado por fragmentos."""
    try:
        chunks = safe_read("datos_dashboard", "orden,contenido")
        chunks.sort(key=lambda row: int(row.get("orden") or 0))
        return "".join(str(row.get("contenido") or "") for row in chunks)
    except Exception:
        return ""


def _market_number(value: Any) -> float:
    text = str(value or "").strip().strip('"').replace(".", "").replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


DASHBOARD_ESTABLISHMENTS = {
    "hospital de curacavi": "Curacaví",
    "centro referencia salud doctor salvador allende g": "CRS S. Allende",
    "hospital de penaflor": "Peñaflor",
    "hospital de melipilla": "Melipilla",
    "hospital dr felix bulnes cerda": "Félix Bulnes",
    "servicio de salud metropolitano occidente": "SSMOCC (Dirección)",
    "hospital de talagante": "Talagante",
    "hospital san juan de dios": "San Juan de Dios",
    "instituto traumatologico": "Inst. Traumatológico",
}


def _decode_csv(data: bytes) -> str:
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        return data.decode("latin-1")


@st.cache_data(show_spinner=False)
def parse_market_package(
    file_name: str, file_bytes: bytes
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Procesa un ZIP o CSV exportado desde Mercado Público."""
    sources: list[tuple[str, bytes]] = []
    if file_name.lower().endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            members = [
                info for info in archive.infolist()
                if not info.is_dir() and info.filename.lower().endswith(".csv")
            ]
            if not members:
                raise ValueError("El ZIP no contiene archivos CSV.")
            if sum(info.file_size for info in members) > 200 * 1024 * 1024:
                raise ValueError("El contenido descomprimido supera el límite de 200 MB.")
            sources = [(info.filename, archive.read(info)) for info in members]
    elif file_name.lower().endswith(".csv"):
        sources = [(file_name, file_bytes)]
    else:
        raise ValueError("Debes seleccionar un archivo ZIP o CSV.")

    output: list[dict[str, Any]] = []
    months: set[str] = set()
    establishments: set[str] = set()

    for source_name, content in sources:
        reader = csv.reader(
            io.StringIO(_decode_csv(content)), delimiter=";", quotechar='"'
        )
        for index, row in enumerate(reader):
            if index == 0 and any(
                str(cell).strip().upper() == "OC" for cell in row
            ):
                continue
            if len(row) < 18:
                continue

            organization = normalize(row[11])
            establishment = DASHBOARD_ESTABLISHMENTS.get(
                organization, str(row[11] or "").strip()
            )
            if not establishment:
                continue

            sent_date = str(row[6] or "").strip()[:10]
            if len(sent_date) >= 7:
                months.add(sent_date[:7])
            establishments.add(establishment)

            record: dict[str, Any] = {
                "e": establishment,
                "oc": str(row[0] or "").strip(),
                "c": str(row[1] or "").strip().upper(),
                "s": str(row[4] or "").strip().upper(),
                "f": sent_date,
                "li": str(row[2] or "").strip()[:24],
                "p": str(row[9] or "").strip()[:120],
                "pr": str(row[13] or "").strip()[:60],
                "rp": str(row[14] or "").strip(),
                "q": round(_market_number(row[15]), 2),
                "u": round(_market_number(row[16]), 2),
                "t": round(_market_number(row[17])),
            }
            currency = str(row[3] or "").strip().upper()
            if currency and currency != "PESO CHILENO":
                record["m"] = (
                    "USD" if "DOLAR" in currency
                    else "UF" if "FOMENTO" in currency
                    else currency
                )
            output.append(record)

    if not output:
        raise ValueError(
            "No se encontraron filas válidas con el formato de Mercado Público."
        )

    stats = {
        "files": len(sources),
        "records": len(output),
        "months": sorted(months),
        "establishments": sorted(establishments),
    }
    return output, stats


def store_dashboard_dataset(
    db: SheetClient,
    rows: list[dict[str, Any]],
    file_name: str,
    stats: dict[str, Any],
) -> None:
    raw = json.dumps(
        rows, ensure_ascii=False, separators=(",", ":")
    ).encode("utf-8")
    compressed = base64.b64encode(gzip.compress(raw, compresslevel=6)).decode("ascii")
    chunk_size = 45000
    chunks = [
        {
            "id": index + 1,
            "orden": index + 1,
            "contenido": compressed[start:start + chunk_size],
        }
        for index, start in enumerate(range(0, len(compressed), chunk_size))
    ]
    db.replace_records("datos_dashboard", chunks)
    months = stats.get("months") or []
    db.table("cargas_mensuales").insert({
        "nombre_archivo": file_name,
        "fecha_carga": datetime.now().isoformat(timespec="seconds"),
        "registros": len(rows),
        "periodo_min": months[0] if months else "",
        "periodo_max": months[-1] if months else "",
        "establecimientos": len(stats.get("establishments") or []),
    }).execute()


def level_name(value: Any) -> str:
    text = normalize(value)
    if "rojo" in text:
        return "Rojo"
    if "amarillo" in text:
        return "Amarillo"
    if "verde" in text:
        return "Verde"
    return str(value or "").strip().title()


DASHBOARD_NAMES = {
    "curacavi": "Curacaví",
    "hospital de curacavi": "Curacaví",
    "hospital dr felix bulnes cerda": "Félix Bulnes",
    "hospital felix bulnes": "Félix Bulnes",
    "felix bulnes": "Félix Bulnes",
    "hospital de melipilla": "Melipilla",
    "hospital san jose melipilla": "Melipilla",
    "hospital melipilla": "Melipilla",
    "melipilla": "Melipilla",
    "hospital de penaflor": "Peñaflor",
    "hospital penaflor": "Peñaflor",
    "penaflor": "Peñaflor",
    "crs salvador allende": "CRS S. Allende",
    "centro de referencia salud occidente salvador allende": "CRS S. Allende",
    "crs s allende": "CRS S. Allende",
    "hospital san juan de dios": "San Juan de Dios",
    "san juan de dios": "San Juan de Dios",
    "ssmocc direccion": "SSMOCC (Dirección)",
    "direccion del servicio metropolitano occidente": "SSMOCC (Dirección)",
    "direccion servicio salud metropolitano occidente": "SSMOCC (Dirección)",
    "instituto traumatologico": "Inst. Traumatológico",
    "instituto traumatologico dr teodoro gebauer": "Inst. Traumatológico",
    "inst traumatologico": "Inst. Traumatológico",
    "hospital de talagante": "Talagante",
    "hospital adalberto steeger talagante": "Talagante",
    "hospital talagante": "Talagante",
    "talagante": "Talagante",
}


def dashboard_name(value: Any) -> str:
    raw = str(value or "").strip()
    return DASHBOARD_NAMES.get(normalize(raw), raw)


# -----------------------------------------------------------------------------
# DATOS PARA EL DASHBOARD NATIVO
# -----------------------------------------------------------------------------
def contracts_for_html(
    contracts: list[dict[str, Any]],
    establishments: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    names = {row.get("id"): row.get("nombre", "") for row in establishments}
    payload: dict[str, dict[str, Any]] = {}

    for row in contracts:
        establishment = dashboard_name(
            row.get("establecimiento")
            or names.get(row.get("establecimiento_id"))
            or ""
        )
        tender = str(row.get("licitacion") or "").strip()
        if not establishment or not tender:
            continue

        payload[f"{establishment}||{tender}"] = {
            "adj": float(row.get("monto_adjudicado") or 0),
            "awardDate": str(row.get("fecha_adjudicacion") or ""),
            "durationMonths": int(row.get("duracion_meses") or 0),
            "renewalLead": int(row.get("anticipacion_renovacion") or 6),
            "manager": str(row.get("responsable") or ""),
            "contractStatus": str(row.get("estado") or "Vigente"),
            "notes": str(row.get("observaciones") or ""),
            "updatedAt": str(row.get("ultima_actualizacion") or ""),
        }
    return payload


def decode_observations(value: Any) -> tuple[str, str]:
    text = str(value or "").strip()
    if not text:
        return "", ""

    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return (
                str(parsed.get("causas") or "").strip(),
                str(parsed.get("medidas") or "").strip(),
            )
    except (ValueError, TypeError):
        pass

    causes_match = re.search(
        r"Principales causas:\s*(.*?)(?=\n\s*\nMedidas implementadas:|$)",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    measures_match = re.search(
        r"Medidas implementadas:\s*(.*)$",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    return (
        causes_match.group(1).strip() if causes_match else "",
        measures_match.group(1).strip() if measures_match else "",
    )


def latest_plan(planes: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not planes:
        return None
    return max(
        planes,
        key=lambda row: (
            str(row.get("fecha_publicacion") or ""),
            int(row.get("id") or 0),
        ),
    )


def plan_for_html(
    planes: list[dict[str, Any]],
    plan_rows: list[dict[str, Any]],
    establishments: list[dict[str, Any]],
) -> dict[str, Any]:
    metadata = latest_plan(planes)
    if metadata is None or not plan_rows:
        return {"meta": None, "items": []}

    names = {row.get("id"): row.get("nombre", "") for row in establishments}
    items: list[dict[str, Any]] = []

    for row in plan_rows:
        establishment = dashboard_name(
            row.get("establecimiento")
            or names.get(row.get("establecimiento_id"))
            or ""
        )
        if not establishment:
            continue

        causes, measures = decode_observations(row.get("observaciones"))
        items.append(
            {
                "estab": establishment,
                "nivel": level_name(row.get("nivel")),
                "periodo": str(metadata.get("periodo") or ""),
                "causas": causes,
                "medidas": measures,
                "compromisos": str(row.get("acciones") or "").strip(),
                "responsable": str(row.get("responsable") or "").strip(),
                "fecha": str(row.get("fecha_compromiso") or "").strip(),
            }
        )

    published_at = metadata.get("creado") or metadata.get("fecha_publicacion")
    timestamp = str(published_at or datetime.now().isoformat())
    if "T" not in timestamp:
        timestamp += "T12:00:00"

    return {
        "meta": {
            "filename": str(metadata.get("nombre_archivo") or "Anexo N°1"),
            "ts": timestamp,
            "reporte": str(metadata.get("reporte") or "Anexo N°1"),
            "periodo": str(metadata.get("periodo") or ""),
        },
        "items": items,
    }


def attach_minsal_results(payload, official_rows, establishments):
    """Agrega las cifras MINSAL, incluso si el Anexo no trae una fila verde."""
    meta = payload.get("meta") or {}
    report_key = normalize(meta.get("reporte"))
    period_key = normalize(meta.get("periodo"))
    names: dict[Any, str] = {}
    for establishment_row in establishments:
        establishment_id = establishment_row.get("id")
        establishment_name = establishment_row.get("nombre", "")
        names[establishment_id] = establishment_name
        names[str(establishment_id)] = establishment_name

    official_by_name: dict[str, dict[str, Any]] = {}
    for row in official_rows:
        if (normalize(row.get("reporte")) != report_key
                or normalize(row.get("periodo")) != period_key):
            continue
        establishment_id = row.get("establecimiento_id")
        establishment = dashboard_name(
            row.get("establecimiento")
            or names.get(establishment_id)
            or names.get(str(establishment_id))
            or DEIS_DASHBOARD_NAMES.get(str(row.get("codigo_deis") or ""))
            or ""
        )
        if establishment:
            official_by_name[establishment] = row

    items = payload.setdefault("items", [])
    items_by_name = {
        dashboard_name(item.get("estab")): item
        for item in items if item.get("estab")
    }
    for establishment, official in official_by_name.items():
        item = items_by_name.get(establishment)
        if item is None:
            item = {
                "estab": establishment,
                "nivel": level_name(official.get("nivel")),
                "periodo": str(meta.get("periodo") or ""),
                "causas": "",
                "medidas": "",
                "compromisos": "",
                "responsable": "",
                "fecha": "",
                "soloResultadoMinsal": True,
            }
            items.append(item)
            items_by_name[establishment] = item
        item["pct"] = _official_percentage(official.get("porcentaje_td"))
        item["minsalNumerador"] = official.get("numerador")
        item["minsalDenominador"] = official.get("denominador")
        item["minsalNivel"] = level_name(official.get("nivel"))
    return payload

# -----------------------------------------------------------------------------
# ADMINISTRACIÓN
# -----------------------------------------------------------------------------
def plan_history_for_html(
    planes: list[dict[str, Any]],
    plan_rows: list[dict[str, Any]],
    establishments: list[dict[str, Any]],
    official_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Construye un plan independiente para cada reporte y período publicado."""
    if not planes:
        return []

    latest_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for metadata in sorted(
        planes,
        key=lambda row: (
            str(row.get("fecha_publicacion") or ""),
            int(row.get("id") or 0),
        ),
    ):
        key = (
            normalize(metadata.get("reporte")),
            normalize(metadata.get("periodo")),
        )
        latest_by_key[key] = metadata

    latest_metadata = latest_plan(planes)
    history: list[dict[str, Any]] = []
    for key, metadata in latest_by_key.items():
        details = [
            row for row in plan_rows
            if (
                normalize(row.get("reporte")),
                normalize(row.get("periodo")),
            ) == key
        ]
        # Compatibilidad con el primer Anexo cargado antes de habilitar historial.
        if metadata is latest_metadata:
            details.extend(
                row for row in plan_rows
                if not row.get("reporte") and not row.get("periodo")
            )
        payload = plan_for_html([metadata], details, establishments)
        payload = attach_minsal_results(payload, official_rows, establishments)
        if payload.get("items"):
            history.append(payload)

    def report_order(payload: dict[str, Any]) -> tuple[int, str]:
        report = str((payload.get("meta") or {}).get("reporte") or "")
        match = re.search(r"(\d+)", report)
        number = int(match.group(1)) if match else 99
        return number, str((payload.get("meta") or {}).get("ts") or "")

    return sorted(history, key=report_order)


def password_hash(password: str) -> str:
    iterations = 210_000
    salt = pysecrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256", password.encode(), bytes.fromhex(salt), iterations
    ).hex()
    return "pbkdf2_sha256$" + str(iterations) + "$" + salt + "$" + digest


def password_matches(password: str, encoded: str) -> bool:
    try:
        algorithm, iterations, salt, expected = str(encoded).split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        digest = hashlib.pbkdf2_hmac(
            "sha256", password.encode(), bytes.fromhex(salt), int(iterations)
        ).hex()
        return hmac.compare_digest(digest, expected)
    except (TypeError, ValueError):
        return False


def same_id(left: Any, right: Any) -> bool:
    return str(left or "").strip() == str(right or "").strip()


def _active_user(value: Any) -> bool:
    return value not in (False, "", 0, "0", "FALSE", "false", None)


def render_user_admin(data, db):
    st.subheader("👥 Usuarios por establecimiento")
    st.info(
        "Cada cuenta queda asociada a un solo establecimiento y no puede "
        "consultar información de otros hospitales."
    )
    establishments = {
        row.get("nombre"): row.get("id")
        for row in data["establecimientos"]
        if row.get("nombre") and row.get("id") is not None
    }
    users = [dict(row) for row in data.get("usuarios_establecimientos", [])]
    id_names = {
        str(row.get("id")): dashboard_name(row.get("nombre"))
        for row in data["establecimientos"]
    }
    if users:
        st.dataframe([{
            "Usuario": row.get("usuario"),
            "Responsable": row.get("nombre"),
            "Establecimiento": id_names.get(
                str(row.get("establecimiento_id")), ""
            ),
            "Estado": "Activo" if _active_user(row.get("activo")) else "Inactivo",
        } for row in users], use_container_width=True, hide_index=True)

    choices = ["Crear nueva cuenta"] + [
        str(row.get("usuario")) for row in users if row.get("usuario")
    ]
    chosen = st.selectbox("Cuenta a gestionar", choices)
    selected = next(
        (row for row in users if row.get("usuario") == chosen), {}
    )
    establishment_names = list(establishments)
    current_name = next((
        name for name, establishment_id in establishments.items()
        if same_id(establishment_id, selected.get("establecimiento_id"))
    ), establishment_names[0] if establishment_names else "")
    with st.form("user_form"):
        username = st.text_input(
            "Usuario", value=str(selected.get("usuario") or ""),
            disabled=bool(selected)
        )
        person = st.text_input(
            "Nombre responsable", value=str(selected.get("nombre") or "")
        )
        establishment = st.selectbox(
            "Establecimiento", establishment_names,
            index=establishment_names.index(current_name)
            if current_name in establishment_names else 0
        )
        password = st.text_input(
            "Contraseña temporal" if not selected
            else "Nueva contraseña (opcional)",
            type="password"
        )
        active = st.checkbox(
            "Cuenta activa", value=_active_user(selected.get("activo", True))
        )
        save = st.form_submit_button(
            "💾 Guardar cuenta", use_container_width=True
        )
    if save:
        clean_user = username.strip().lower()
        if len(clean_user) < 4:
            st.error("El usuario debe tener al menos 4 caracteres.")
            return
        if (not selected and any(
            normalize(row.get("usuario")) == normalize(clean_user)
            for row in users
        )):
            st.error("Ese usuario ya existe.")
            return
        if (not selected and len(password) < 8) or (
            password and len(password) < 8
        ):
            st.error("La contraseña debe tener al menos 8 caracteres.")
            return
        try:
            user_table = db.table("usuarios_establecimientos")
            if selected:
                payload = {
                    "nombre": person.strip(),
                    "establecimiento_id": establishments[establishment],
                    "activo": active,
                }
                if password:
                    payload["clave_hash"] = password_hash(password)
                user_table.update(payload).eq(
                    "id", selected.get("id")
                ).execute()
            else:
                next_id = max(
                    [int(row.get("id") or 0) for row in users] + [0]
                ) + 1
                user_table.insert({
                    "id": next_id,
                    "usuario": clean_user,
                    "clave_hash": password_hash(password),
                    "establecimiento_id": establishments[establishment],
                    "nombre": person.strip(),
                    "activo": active,
                    "creado": datetime.now().isoformat(timespec="seconds"),
                    "ultima_conexion": "",
                }).execute()
            st.success(
                f"Cuenta {clean_user} guardada. Ya puede ingresar desde "
                "Portal establecimientos."
            )
            st.cache_data.clear()
            st.rerun()
        except Exception as exc:
            st.error(
                "No fue posible guardar la cuenta en Google Sheets. "
                f"Detalle: {exc}"
            )


def _contract_date(value: Any) -> str:
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return ""


@st.cache_data(show_spinner=False)
def parse_contract_bulk(file_name, file_bytes, establishments, contracts):
    text = _decode_csv(file_bytes)
    delimiter = ";" if text[:4096].count(";") >= text[:4096].count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = {
        normalize(header): header for header in (reader.fieldnames or [])
    }
    if "establecimiento" not in headers or "instrumento" not in headers:
        raise ValueError("Faltan las columnas Establecimiento e Instrumento.")
    establishment_ids = {
        dashboard_name(row.get("nombre")): row.get("id")
        for row in establishments if row.get("id") is not None
    }
    merged = [dict(row) for row in contracts]
    by_key = {
        (str(row.get("establecimiento_id") or ""), normalize(row.get("licitacion"))): row
        for row in merged if row.get("licitacion")
    }
    next_id = max([int(row.get("id") or 0) for row in merged] + [0]) + 1
    created = updated = skipped = 0
    field_map = {
        "monto_adjudicado": "monto adjudicado",
        "fecha_adjudicacion": "fecha adjudicacion",
        "duracion_meses": "duracion meses",
        "anticipacion_renovacion": "anticipacion renovacion meses",
        "responsable": "responsable", "estado": "estado contractual",
        "observaciones": "observaciones",
    }
    for source in reader:
        establishment_id = establishment_ids.get(
            dashboard_name(source.get(headers["establecimiento"], ""))
        )
        instrument = str(source.get(headers["instrumento"], "") or "").strip()
        if establishment_id is None or not instrument:
            skipped += 1
            continue
        key = (str(establishment_id), normalize(instrument))
        target = by_key.get(key)
        if target is None:
            target = {
                "id": next_id, "establecimiento_id": establishment_id,
                "licitacion": instrument, "monto_adjudicado": 0,
                "fecha_adjudicacion": "", "duracion_meses": "",
                "anticipacion_renovacion": 6, "estado": "Vigente",
                "responsable": "", "observaciones": "",
                "estado_revision": "Incompleto",
                "enviado_revision": "", "actualizado_por": "carga_masiva",
            }
            next_id += 1
            merged.append(target)
            by_key[key] = target
            created += 1
        else:
            updated += 1
        for destination, source_name in field_map.items():
            header = headers.get(source_name)
            raw = str(source.get(header, "") or "").strip() if header else ""
            if not raw:
                continue
            if destination in (
                "monto_adjudicado", "duracion_meses",
                "anticipacion_renovacion"
            ):
                target[destination] = int(round(_official_number(raw)))
            elif destination == "fecha_adjudicacion":
                parsed = _contract_date(raw)
                if parsed:
                    target[destination] = parsed
            else:
                target[destination] = raw
        target["ultima_actualizacion"] = datetime.now().isoformat(
            timespec="seconds"
        )
        target["actualizado_por"] = "carga_masiva"
    return merged, {"created": created, "updated": updated, "skipped": skipped}


def render_contract_bulk_admin(data, db):
    st.subheader("📥 Gestión centralizada de antecedentes contractuales")
    st.info(
        "Seleccione un establecimiento, descargue la plantilla con los "
        "antecedentes pendientes y, cuando el hospital la devuelva, "
        "incorpórela desde esta misma sección."
    )

    establishments = [
        row for row in data.get("establecimientos", [])
        if row.get("id") is not None and row.get("nombre")
    ]
    if not establishments:
        st.error("No existen establecimientos configurados.")
        return

    establishment_options = {
        dashboard_name(row.get("nombre")): row
        for row in establishments
    }
    selected_name = st.selectbox(
        "Establecimiento a gestionar",
        list(establishment_options),
        help=(
            "La plantilla y la carga quedarán restringidas al "
            "establecimiento seleccionado."
        ),
    )
    selected_establishment = establishment_options[selected_name]
    establishment_id = selected_establishment.get("id")
    contracts = portal_contracts_for_establishment(
        data, establishment_id, selected_name
    )
    missing_contracts = [
        row for row in contracts
        if not row.get("monto_adjudicado")
        or not row.get("fecha_adjudicacion")
        or not row.get("duracion_meses")
    ]
    completed = len(contracts) - len(missing_contracts)

    c1, c2, c3 = st.columns(3)
    c1.metric("Instrumentos detectados", len(contracts))
    c2.metric("Antecedentes completos", completed)
    c3.metric("Pendientes por solicitar", len(missing_contracts))

    if contracts:
        st.progress(
            completed / len(contracts),
            text=(
                f"Avance de {selected_name}: "
                f"{completed} de {len(contracts)} completos"
            ),
        )

    st.markdown("### 1. Descargar solicitud para el hospital")
    if not missing_contracts:
        st.success(
            "Este establecimiento no tiene antecedentes contractuales "
            "pendientes en la base actual."
        )
    else:
        st.caption(
            "La plantilla contiene únicamente instrumentos donde falta "
            "monto adjudicado, fecha de adjudicación o duración."
        )
        template = portal_template_xlsx(
            selected_name, missing_contracts
        )
        st.download_button(
            "⬇️ Descargar plantilla Excel de antecedentes pendientes",
            data=template,
            file_name=(
                "Pendientes_Contractuales_"
                + re.sub(
                    r"[^A-Za-z0-9]+", "_", normalize(selected_name)
                )
                + ".xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
            type="primary",
        )
        preview = [{
            "Instrumento": row.get("licitacion"),
            "OC": int(row.get("cantidad_oc") or 0),
            "Monto ejecutado": float(row.get("monto_ejecutado") or 0),
            "Falta monto": "Sí" if not row.get("monto_adjudicado") else "No",
            "Falta fecha": "Sí" if not row.get("fecha_adjudicacion") else "No",
            "Falta duración": "Sí" if not row.get("duracion_meses") else "No",
        } for row in missing_contracts]
        with st.expander(
            f"Ver los {len(missing_contracts)} antecedentes solicitados"
        ):
            st.dataframe(
                preview,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Monto ejecutado": st.column_config.NumberColumn(
                        format="$ %d"
                    )
                },
            )

    st.markdown("### 2. Incorporar respuesta del hospital")
    uploaded = st.file_uploader(
        "Seleccionar plantilla Excel completada",
        type=["xlsx"],
        key=f"admin_contract_bulk_{establishment_id}",
        help=(
            "Utilice la misma plantilla descargada desde esta sección. "
            "El sistema no aceptará instrumentos de otro hospital."
        ),
    )
    review_status = st.selectbox(
        "Estado después de incorporar",
        ["Borrador", "Validado"],
        help=(
            "Borrador permite revisar posteriormente. Validado bloquea "
            "la edición del establecimiento."
        ),
    )

    records: list[dict[str, Any]] = []
    omitted = 0
    if uploaded:
        try:
            frame = pd.read_excel(uploaded, sheet_name="Contratos")
            frame.columns = [
                str(column).strip() for column in frame.columns
            ]
            required = {
                "Instrumento", "Monto adjudicado",
                "Fecha adjudicación", "Duración meses",
            }
            missing_columns = sorted(required - set(frame.columns))
            if missing_columns:
                raise ValueError(
                    "Faltan columnas: " + ", ".join(missing_columns)
                )

            allowed = {
                normalize(row.get("licitacion")) for row in contracts
            }
            for source in frame.to_dict("records"):
                if normalize(source.get("Instrumento")) not in allowed:
                    omitted += 1
                    continue
                record = _portal_record_from_row(
                    source, establishment_id, "administrador"
                )
                if record:
                    record["estado_revision"] = review_status
                    records.append(record)
                else:
                    omitted += 1

            v1, v2, v3 = st.columns(3)
            v1.metric("Filas del archivo", len(frame))
            v2.metric("Listas para incorporar", len(records))
            v3.metric("Incompletas u omitidas", omitted)

            if records:
                st.success(
                    "Archivo validado. Revise la vista previa antes "
                    "de incorporar."
                )
                preview_frame = pd.DataFrame(records)[[
                    "licitacion", "monto_adjudicado",
                    "fecha_adjudicacion", "duracion_meses",
                    "responsable", "estado_revision",
                ]]
                st.dataframe(
                    preview_frame.head(200),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "monto_adjudicado":
                            st.column_config.NumberColumn(format="$ %d")
                    },
                )
        except Exception as exc:
            st.error(f"No fue posible validar la plantilla: {exc}")
            records = []

    confirm = st.checkbox(
        "Confirmo que revisé la vista previa y deseo actualizar "
        f"únicamente {selected_name}.",
        disabled=not records,
    )
    if st.button(
        "📥 Incorporar antecedentes validados",
        type="primary",
        use_container_width=True,
        disabled=not records or not confirm,
    ):
        try:
            upsert_portal_contracts(db, records)
            db.table("historial_cambios").insert({
                "contrato_id": "",
                "establecimiento_id": establishment_id,
                "usuario": "administrador",
                "accion": "Carga centralizada",
                "detalle": (
                    f"{len(records)} instrumentos incorporados para "
                    f"{selected_name} con estado {review_status}"
                ),
                "fecha": datetime.now().isoformat(timespec="seconds"),
            }).execute()
            st.success(
                f"Carga completada: {len(records)} instrumentos "
                f"actualizados para {selected_name}."
            )
            st.cache_data.clear()
            st.rerun()
        except Exception as exc:
            st.error(f"No fue posible incorporar la carga: {exc}")


@st.cache_data(show_spinner=False)
def decode_dashboard_dataset(dataset_gzip_b64: str) -> list[dict[str, Any]]:
    """Decodifica una vez la última base mensual guardada en Google Sheets."""
    if not dataset_gzip_b64:
        return []
    try:
        compressed = base64.b64decode(dataset_gzip_b64)
        raw = gzip.decompress(compressed)
        records = json.loads(raw.decode("utf-8"))
        return records if isinstance(records, list) else []
    except (ValueError, OSError, json.JSONDecodeError):
        return []


def portal_contracts_for_establishment(
    data: dict[str, Any],
    establishment_id: Any,
    establishment_name: str,
) -> list[dict[str, Any]]:
    """Homologa instrumentos mensuales con antecedentes contractuales guardados."""
    saved = [
        dict(row) for row in data.get("contratos", [])
        if same_id(row.get("establecimiento_id"), establishment_id)
    ]
    by_tender = {
        normalize(row.get("licitacion")): row
        for row in saved if row.get("licitacion")
    }
    monthly_rows = decode_dashboard_dataset(
        str(data.get("dataset_gzip_b64") or "")
    )
    target_name = dashboard_name(establishment_name)
    aggregates: dict[str, dict[str, Any]] = {}
    for row in monthly_rows:
        if dashboard_name(row.get("e")) != target_name:
            continue
        tender = str(row.get("li") or "").strip()
        if not tender:
            continue
        key = normalize(tender)
        item = aggregates.setdefault(key, {
            "licitacion": tender,
            "monto_ejecutado": 0.0,
            "ordenes_compra": set(),
        })
        try:
            item["monto_ejecutado"] += float(row.get("t") or 0)
        except (TypeError, ValueError):
            pass
        purchase_order = str(row.get("oc") or "").strip()
        if purchase_order:
            item["ordenes_compra"].add(purchase_order)

    for key, source in aggregates.items():
        if key in by_tender:
            by_tender[key]["monto_ejecutado"] = round(
                source["monto_ejecutado"]
            )
            by_tender[key]["cantidad_oc"] = len(source["ordenes_compra"])
            continue
        by_tender[key] = {
            "id": None,
            "establecimiento_id": establishment_id,
            "licitacion": source["licitacion"],
            "monto_adjudicado": 0,
            "fecha_adjudicacion": "",
            "duracion_meses": "",
            "anticipacion_renovacion": 6,
            "estado": "Vigente",
            "responsable": "",
            "observaciones": "",
            "estado_revision": "Incompleto",
            "monto_ejecutado": round(source["monto_ejecutado"]),
            "cantidad_oc": len(source["ordenes_compra"]),
            "_origen_mensual": True,
        }
    return sorted(
        by_tender.values(),
        key=lambda row: normalize(row.get("licitacion")),
    )


def _portal_contract_key(row: dict[str, Any]) -> tuple[str, str]:
    return (
        str(row.get("establecimiento_id") or "").strip(),
        normalize(row.get("licitacion")),
    )


def upsert_portal_contracts(
    db: SheetClient, records: list[dict[str, Any]]
) -> None:
    """Fusiona una edición múltiple con la versión más reciente de la hoja."""
    current = [
        dict(row)
        for row in db.table("contratos").select("*").execute().data
    ]
    by_key = {
        _portal_contract_key(row): row
        for row in current if row.get("licitacion")
    }
    next_id = max(
        [int(row.get("id") or 0) for row in current] + [0]
    ) + 1
    for incoming in records:
        key = _portal_contract_key(incoming)
        target = by_key.get(key)
        if target is None:
            target = {"id": next_id}
            next_id += 1
            current.append(target)
            by_key[key] = target
        target.update(incoming)
    db.replace_records("contratos", current)


@st.cache_data(show_spinner=False)
def portal_template_xlsx(
    establishment_name: str,
    contracts: list[dict[str, Any]],
) -> bytes:
    """Crea una plantilla Excel profesional y precargada para el hospital."""
    columns = [
        "Establecimiento", "Instrumento", "OC", "Monto ejecutado",
        "Monto adjudicado", "Fecha adjudicación", "Duración meses",
        "Anticipación renovación meses", "Estado contractual",
        "Responsable", "Observaciones",
    ]
    rows = []
    for row in contracts:
        rows.append({
            "Establecimiento": establishment_name,
            "Instrumento": row.get("licitacion") or "",
            "OC": int(row.get("cantidad_oc") or 0),
            "Monto ejecutado": float(row.get("monto_ejecutado") or 0),
            "Monto adjudicado": float(row.get("monto_adjudicado") or 0),
            "Fecha adjudicación": row.get("fecha_adjudicacion") or "",
            "Duración meses": row.get("duracion_meses") or "",
            "Anticipación renovación meses":
                row.get("anticipacion_renovacion") or 6,
            "Estado contractual": row.get("estado") or "Vigente",
            "Responsable": row.get("responsable") or "",
            "Observaciones": row.get("observaciones") or "",
        })
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame = pd.DataFrame(rows, columns=columns)
        frame.to_excel(writer, sheet_name="Contratos", index=False)
        workbook = writer.book
        instructions = workbook.create_sheet("Instrucciones", 0)
        instructions["A1"] = "PLANTILLA DE ANTECEDENTES CONTRACTUALES"
        instructions["A2"] = establishment_name
        instructions["A4"] = "Cómo completar"
        instructions["A5"] = (
            "1. No modifique Establecimiento, Instrumento, OC ni Monto ejecutado."
        )
        instructions["A6"] = (
            "2. Complete Monto adjudicado, Fecha adjudicación y Duración meses."
        )
        instructions["A7"] = (
            "3. Use fechas con formato DD/MM/AAAA o AAAA-MM-DD."
        )
        instructions["A8"] = (
            "4. Puede completar responsable, estado y observaciones."
        )
        instructions["A9"] = (
            "5. Guarde el archivo y súbalo en Carga masiva del portal."
        )
        instructions["A11"] = (
            "Seguridad: el portal solo aceptará instrumentos asociados "
            "a este establecimiento."
        )

        from openpyxl.styles import Alignment, Font, PatternFill
        navy = "003B6F"
        blue = "0B6FB8"
        light_blue = "DCEAF7"
        gray = "E7ECF2"
        white = "FFFFFF"

        instructions["A1"].font = Font(
            bold=True, size=16, color=white
        )
        instructions["A1"].fill = PatternFill("solid", fgColor=navy)
        instructions["A2"].font = Font(
            bold=True, size=13, color=navy
        )
        instructions["A4"].font = Font(
            bold=True, size=12, color=blue
        )
        instructions.column_dimensions["A"].width = 95
        for row_number in range(5, 12):
            instructions[f"A{row_number}"].alignment = Alignment(wrap_text=True)
        instructions.freeze_panes = "A4"

        sheet = writer.sheets["Contratos"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = [25, 24, 10, 18, 18, 18, 16, 24, 20, 24, 45]
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[
                chr(64 + index) if index <= 26 else "A"
            ].width = width
        for cell in sheet[1]:
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for row_number in range(2, sheet.max_row + 1):
            for column_number in (1, 2, 3, 4):
                sheet.cell(row_number, column_number).fill = PatternFill(
                    "solid", fgColor=gray
                )
            for column_number in range(5, 12):
                sheet.cell(row_number, column_number).fill = PatternFill(
                    "solid", fgColor=light_blue
                )
            sheet.cell(row_number, 4).number_format = '"$"#,##0'
            sheet.cell(row_number, 5).number_format = '"$"#,##0'
        sheet.sheet_view.showGridLines = False
        sheet.row_dimensions[1].height = 32
    return output.getvalue()


def _portal_date(value: Any) -> str:
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _contract_date(value)


def _portal_record_from_row(
    row: dict[str, Any],
    establishment_id: Any,
    username: str,
) -> dict[str, Any] | None:
    tender = str(row.get("Instrumento") or "").strip()
    amount = _official_number(row.get("Monto adjudicado"))
    award_date = _portal_date(row.get("Fecha adjudicación"))
    duration = int(round(_official_number(row.get("Duración meses"))))
    if not tender or amount <= 0 or not award_date or duration <= 0:
        return None
    renewal = int(round(_official_number(
        row.get("Anticipación renovación meses") or 6
    )))
    return {
        "establecimiento_id": establishment_id,
        "licitacion": tender,
        "monto_adjudicado": amount,
        "fecha_adjudicacion": award_date,
        "duracion_meses": duration,
        "anticipacion_renovacion": max(0, renewal),
        "estado": str(row.get("Estado contractual") or "Vigente").strip(),
        "responsable": str(row.get("Responsable") or "").strip(),
        "observaciones": str(row.get("Observaciones") or "").strip(),
        "ultima_actualizacion": datetime.now().isoformat(timespec="seconds"),
        "estado_revision": "Borrador",
        "actualizado_por": username,
    }


def render_portal_quick_editor(
    data: dict[str, Any],
    db: SheetClient,
    contracts: list[dict[str, Any]],
    establishment_id: Any,
    username: str,
) -> None:
    st.subheader("Edición rápida en línea")
    st.caption(
        "Busque instrumentos y complete varias filas directamente. "
        "Solo se guardarán filas con monto, fecha y duración válidos."
    )
    left, middle, right = st.columns([3, 1, 1])
    search = left.text_input(
        "Buscar instrumento", placeholder="Ej.: 1641-121-LR24"
    )
    pending_only = middle.checkbox("Solo pendientes", value=True)
    page_size = right.selectbox("Filas por página", [25, 50, 100], index=1)

    filtered = []
    for row in contracts:
        if search and normalize(search) not in normalize(row.get("licitacion")):
            continue
        complete = bool(
            row.get("monto_adjudicado")
            and row.get("fecha_adjudicacion")
            and row.get("duracion_meses")
        )
        if pending_only and complete:
            continue
        filtered.append(row)

    pages = max(1, (len(filtered) + page_size - 1) // page_size)
    page = st.number_input(
        "Página", min_value=1, max_value=pages, value=1, step=1
    )
    start = (int(page) - 1) * page_size
    visible = filtered[start:start + page_size]
    st.caption(
        f"Mostrando {start + 1 if visible else 0}–"
        f"{start + len(visible)} de {len(filtered)} instrumentos."
    )
    editable_rows = [{
        "Instrumento": row.get("licitacion") or "",
        "OC": int(row.get("cantidad_oc") or 0),
        "Monto ejecutado": float(row.get("monto_ejecutado") or 0),
        "Monto adjudicado": float(row.get("monto_adjudicado") or 0),
        "Fecha adjudicación": row.get("fecha_adjudicacion") or "",
        "Duración meses": int(row.get("duracion_meses") or 0),
        "Anticipación renovación meses": int(
            row.get("anticipacion_renovacion") or 6
        ),
        "Estado contractual": row.get("estado") or "Vigente",
        "Responsable": row.get("responsable") or "",
        "Observaciones": row.get("observaciones") or "",
    } for row in visible]
    edited = st.data_editor(
        pd.DataFrame(editable_rows),
        use_container_width=True,
        hide_index=True,
        disabled=["Instrumento", "OC", "Monto ejecutado"],
        num_rows="fixed",
        column_config={
            "Monto ejecutado": st.column_config.NumberColumn(format="$ %d"),
            "Monto adjudicado": st.column_config.NumberColumn(
                format="$ %d", min_value=0
            ),
            "Fecha adjudicación": st.column_config.TextColumn(
                help="DD/MM/AAAA o AAAA-MM-DD"
            ),
            "Duración meses": st.column_config.NumberColumn(min_value=0),
            "Observaciones": st.column_config.TextColumn(width="large"),
        },
        key=f"portal_quick_{page}_{pending_only}_{search}",
    )
    if st.button(
        "💾 Guardar filas completas como borrador",
        type="primary",
        use_container_width=True,
        disabled=edited.empty,
    ):
        allowed = {
            normalize(row.get("licitacion")) for row in contracts
        }
        records = []
        for source in edited.to_dict("records"):
            if normalize(source.get("Instrumento")) not in allowed:
                continue
            record = _portal_record_from_row(
                source, establishment_id, username
            )
            if record:
                records.append(record)
        if not records:
            st.error(
                "No hay filas completas. Ingrese monto adjudicado, fecha "
                "y duración antes de guardar."
            )
            return
        try:
            upsert_portal_contracts(db, records)
            db.table("historial_cambios").insert({
                "contrato_id": "",
                "establecimiento_id": establishment_id,
                "usuario": username,
                "accion": "Edición rápida",
                "detalle": f"{len(records)} instrumentos guardados como borrador",
                "fecha": datetime.now().isoformat(timespec="seconds"),
            }).execute()
            st.success(f"Se guardaron {len(records)} instrumentos.")
            st.cache_data.clear()
            st.rerun()
        except Exception as exc:
            st.error(f"No fue posible guardar la edición rápida: {exc}")


def render_portal_bulk_upload(
    data: dict[str, Any],
    db: SheetClient,
    contracts: list[dict[str, Any]],
    establishment_id: Any,
    establishment_name: str,
    username: str,
) -> None:
    st.subheader("Carga masiva con plantilla")
    st.info(
        "Descargue la plantilla precargada, complete las columnas azules "
        "y vuelva a subir el archivo. Los datos quedarán como borrador."
    )
    template = portal_template_xlsx(establishment_name, contracts)
    st.download_button(
        "⬇️ Descargar plantilla Excel del establecimiento",
        data=template,
        file_name=(
            "Plantilla_Contratos_"
            + re.sub(r"[^A-Za-z0-9]+", "_", normalize(establishment_name))
            + ".xlsx"
        ),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
    )
    uploaded = st.file_uploader(
        "Subir plantilla completada",
        type=["xlsx"],
        key="portal_contract_bulk",
    )
    records: list[dict[str, Any]] = []
    omitted = 0
    if uploaded:
        try:
            frame = pd.read_excel(uploaded, sheet_name="Contratos")
            frame.columns = [str(column).strip() for column in frame.columns]
            required = {
                "Instrumento", "Monto adjudicado",
                "Fecha adjudicación", "Duración meses",
            }
            missing = sorted(required - set(frame.columns))
            if missing:
                raise ValueError("Faltan columnas: " + ", ".join(missing))
            allowed = {
                normalize(row.get("licitacion")) for row in contracts
            }
            for source in frame.to_dict("records"):
                if normalize(source.get("Instrumento")) not in allowed:
                    omitted += 1
                    continue
                record = _portal_record_from_row(
                    source, establishment_id, username
                )
                if record:
                    records.append(record)
                else:
                    omitted += 1
            a, b = st.columns(2)
            a.metric("Filas listas para guardar", len(records))
            b.metric("Filas incompletas u omitidas", omitted)
            if records:
                st.dataframe(
                    pd.DataFrame(records)[[
                        "licitacion", "monto_adjudicado",
                        "fecha_adjudicacion", "duracion_meses",
                        "responsable",
                    ]].head(100),
                    use_container_width=True,
                    hide_index=True,
                )
        except Exception as exc:
            st.error(f"No fue posible validar la plantilla: {exc}")
            records = []

    if st.button(
        "📥 Guardar carga masiva como borrador",
        type="primary",
        use_container_width=True,
        disabled=not records,
    ):
        try:
            upsert_portal_contracts(db, records)
            db.table("historial_cambios").insert({
                "contrato_id": "",
                "establecimiento_id": establishment_id,
                "usuario": username,
                "accion": "Carga masiva",
                "detalle": f"{len(records)} instrumentos guardados como borrador",
                "fecha": datetime.now().isoformat(timespec="seconds"),
            }).execute()
            st.success(
                f"Carga finalizada: {len(records)} instrumentos guardados."
            )
            st.cache_data.clear()
            st.rerun()
        except Exception as exc:
            st.error(f"No fue posible guardar la carga masiva: {exc}")

def portal_login(data):
    user_id = st.session_state.get("portal_user_id")
    if user_id is not None:
        current = next((
            row for row in data.get("usuarios_establecimientos", [])
            if same_id(row.get("id"), user_id)
            and _active_user(row.get("activo"))
        ), None)
        if current:
            return current
        st.session_state.pop("portal_user_id", None)

    st.markdown("## 🏥 Portal de establecimientos")
    st.caption("Complete y remita antecedentes contractuales al SSMOCC.")
    with st.form("portal_login"):
        username = st.text_input("Usuario")
        password = st.text_input("Contraseña", type="password")
        submit = st.form_submit_button(
            "Ingresar al portal", use_container_width=True
        )
    if submit:
        current = next((
            row for row in data.get("usuarios_establecimientos", [])
            if normalize(row.get("usuario")) == normalize(username)
            and _active_user(row.get("activo"))
        ), None)
        if current and password_matches(
            password, str(current.get("clave_hash") or "")
        ):
            st.session_state["portal_user_id"] = current.get("id")
            st.rerun()
        st.error("Usuario o contraseña incorrectos.")
    if st.button("← Volver al dashboard"):
        st.query_params.clear()
        st.rerun()
    return None


def render_establishment_portal(data, db):
    user = portal_login(data)
    if not user:
        return

    # Mantiene el portal en la misma página durante todos los reruns.
    st.query_params["portal"] = "1"
    establishment_id = user.get("establecimiento_id")
    establishment = next((
        row for row in data["establecimientos"]
        if same_id(row.get("id"), establishment_id)
    ), {})
    name = dashboard_name(establishment.get("nombre") or "Establecimiento")

    left, right = st.columns([6, 1])
    left.success(
        f"Sesión activa · {name} · {user.get('nombre') or user.get('usuario')}"
    )
    if right.button("Cerrar sesión"):
        st.session_state.pop("portal_user_id", None)
        st.query_params.clear()
        st.rerun()

    st.title("Portal de gestión contractual")
    st.caption(
        "Complete los antecedentes de sus instrumentos. Este portal muestra "
        "exclusivamente los registros asociados a su establecimiento."
    )

    contracts = portal_contracts_for_establishment(
        data, establishment_id, name
    )
    incomplete = sum(
        1 for row in contracts
        if not row.get("monto_adjudicado")
        or not row.get("fecha_adjudicacion")
        or not row.get("duracion_meses")
    )
    in_review = sum(
        normalize(row.get("estado_revision")) == "enviado"
        for row in contracts
    )
    validated = sum(
        normalize(row.get("estado_revision")) == "validado"
        for row in contracts
    )
    completed = max(0, len(contracts) - incomplete)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Instrumentos asignados", len(contracts))
    c2.metric("Pendientes de completar", incomplete)
    c3.metric("En revisión SSMOCC", in_review)
    c4.metric("Validados", validated)

    if not contracts:
        st.info(
            "El administrador todavía no ha asignado instrumentos a este "
            "establecimiento."
        )
        return

    progress = completed / len(contracts) if contracts else 0
    st.progress(
        progress,
        text=f"Avance de antecedentes completos: {completed} de {len(contracts)}",
    )

    work_mode = st.radio(
        "Forma de trabajo",
        ["✏️ Edición rápida", "📥 Carga masiva", "🔎 Detalle individual"],
        horizontal=True,
        help=(
            "Puede completar datos directamente, usar una plantilla Excel "
            "o revisar un instrumento en detalle."
        ),
    )
    username = str(user.get("usuario") or "")
    if work_mode == "✏️ Edición rápida":
        render_portal_quick_editor(
            data, db, contracts, establishment_id, username
        )
        return
    if work_mode == "📥 Carga masiva":
        render_portal_bulk_upload(
            data, db, contracts, establishment_id, name, username
        )
        return

    summary_rows = []
    for row in contracts:
        complete = bool(
            row.get("monto_adjudicado")
            and row.get("fecha_adjudicacion")
            and row.get("duracion_meses")
        )
        summary_rows.append({
            "Instrumento": row.get("licitacion") or "Sin código",
            "OC": int(row.get("cantidad_oc") or 0),
            "Monto ejecutado": float(row.get("monto_ejecutado") or 0),
            "Antecedentes": "Completos" if complete else "Pendientes",
            "Revisión SSMOCC": row.get("estado_revision") or "Incompleto",
            "Última actualización": row.get("ultima_actualizacion") or "—",
        })
    st.dataframe(
        summary_rows, use_container_width=True, hide_index=True,
        column_config={
            "Instrumento": st.column_config.TextColumn(width="medium"),
            "OC": st.column_config.NumberColumn(width="small"),
            "Monto ejecutado": st.column_config.NumberColumn(
                format="$ %d", width="medium"
            ),
            "Antecedentes": st.column_config.TextColumn(width="small"),
            "Revisión SSMOCC": st.column_config.TextColumn(width="small"),
        },
    )

    st.subheader("Completar o actualizar un instrumento")
    st.info(
        "Seleccione una licitación, complete monto adjudicado, fecha y "
        "duración; luego guarde un borrador o envíelo a revisión."
    )
    options = {
        (
            f"{row.get('licitacion') or 'Sin código'} · "
            f"{row.get('estado_revision') or 'Incompleto'}"
        ): row
        for row in contracts
    }
    selected_label = st.selectbox(
        "Instrumento a gestionar", list(options),
        help="Solo aparecen instrumentos pertenecientes a su establecimiento.",
    )
    selected = options[selected_label]
    locked = normalize(selected.get("estado_revision")) == "validado"

    raw_date = str(selected.get("fecha_adjudicacion") or "").strip()
    try:
        default_date = date.fromisoformat(raw_date)
    except (TypeError, ValueError):
        default_date = date.today()

    review_status = selected.get("estado_revision") or "Incompleto"
    if locked:
        st.success(
            "Este instrumento fue validado por el SSMOCC y quedó bloqueado "
            "para proteger la información oficial."
        )
    elif normalize(review_status) == "enviado":
        st.warning(
            "Este instrumento está en revisión. Puede actualizarlo y volver "
            "a enviarlo mientras no haya sido validado."
        )

    with st.form("portal_contract_form", clear_on_submit=False):
        st.text_input(
            "Licitación / instrumento",
            value=str(selected.get("licitacion") or ""),
            disabled=True,
        )
        first, second, third = st.columns([1.3, 1, 1])
        amount = first.number_input(
            "Monto adjudicado (CLP)",
            min_value=0.0,
            step=100000.0,
            value=float(selected.get("monto_adjudicado") or 0),
            format="%.0f",
            disabled=locked,
            help="Ingrese el monto total adjudicado del instrumento.",
        )
        award_date = second.date_input(
            "Fecha de adjudicación",
            value=default_date,
            format="DD/MM/YYYY",
            disabled=locked,
        )
        duration = third.number_input(
            "Duración del contrato (meses)",
            min_value=1,
            max_value=240,
            value=int(selected.get("duracion_meses") or 12),
            disabled=locked,
        )

        col1, col2 = st.columns(2)
        renewal = col1.number_input(
            "Anticipación de renovación (meses)",
            min_value=0,
            max_value=36,
            value=int(selected.get("anticipacion_renovacion") or 6),
            disabled=locked,
        )
        statuses = [
            "Vigente", "En renovación", "Prorrogado",
            "Finalizado", "Suspendido",
        ]
        current_status = str(selected.get("estado") or "Vigente")
        status = col2.selectbox(
            "Estado contractual",
            statuses,
            index=statuses.index(current_status)
            if current_status in statuses else 0,
            disabled=locked,
        )
        manager = st.text_input(
            "Responsable del contrato",
            value=str(selected.get("responsable") or ""),
            disabled=locked,
        )
        observations = st.text_area(
            "Observaciones",
            value=str(selected.get("observaciones") or ""),
            disabled=locked,
            placeholder="Registre antecedentes, alertas o comentarios relevantes.",
        )
        b1, b2 = st.columns(2)
        draft = b1.form_submit_button(
            "💾 Guardar borrador",
            use_container_width=True,
            disabled=locked,
        )
        send = b2.form_submit_button(
            "📨 Enviar a revisión SSMOCC",
            use_container_width=True,
            type="primary",
            disabled=locked,
        )

    if draft or send:
        if amount <= 0:
            st.error("Debe ingresar un monto adjudicado mayor que cero.")
            return
        if int(duration) <= 0:
            st.error("Debe ingresar la duración del contrato.")
            return
        if send and not manager.strip():
            st.error(
                "Para enviar a revisión debe indicar el responsable del contrato."
            )
            return

        now = datetime.now().isoformat(timespec="seconds")
        payload = {
            "monto_adjudicado": amount,
            "fecha_adjudicacion": award_date.isoformat(),
            "duracion_meses": int(duration),
            "anticipacion_renovacion": int(renewal),
            "estado": status,
            "responsable": manager.strip(),
            "observaciones": observations.strip(),
            "ultima_actualizacion": now,
            "estado_revision": "Enviado" if send else "Borrador",
            "enviado_revision": now if send else str(
                selected.get("enviado_revision") or ""
            ),
            "actualizado_por": str(user.get("usuario") or ""),
        }
        history_payload = {
            "contrato_id": selected.get("id"),
            "establecimiento_id": establishment_id,
            "usuario": user.get("usuario"),
            "accion": "Enviado" if send else "Borrador",
            "detalle": json.dumps(payload, ensure_ascii=False),
            "fecha": now,
        }
        try:
            if selected.get("id") is not None:
                db.table("contratos").update(payload).eq(
                    "id", selected.get("id")
                ).execute()
            else:
                db.table("contratos").insert({
                    "establecimiento_id": establishment_id,
                    "licitacion": selected.get("licitacion"),
                    **payload,
                }).execute()
            db.table("historial_cambios").insert(history_payload).execute()
            st.success(
                "Antecedentes enviados correctamente al SSMOCC."
                if send
                else "Borrador guardado correctamente."
            )
            st.cache_data.clear()
            st.rerun()
        except Exception as exc:
            st.error(
                "No fue posible guardar los antecedentes en Google Sheets. "
                f"Detalle: {exc}"
            )

    if st.button("← Volver al dashboard público"):
        st.session_state.pop("portal_user_id", None)
        st.query_params.clear()
        st.rerun()

def admin_login() -> bool:
    configured = secret("ADMIN_PASSWORD")
    if not configured:
        st.error("Falta configurar ADMIN_PASSWORD en Streamlit Secrets.")
        return False
    if st.session_state.get("admin_authenticated"):
        return True

    st.markdown("## 🔐 Administración")
    with st.form("admin_login"):
        password = st.text_input("Contraseña administrativa", type="password")
        submitted = st.form_submit_button("Ingresar")
    if submitted:
        if password == configured:
            st.session_state["admin_authenticated"] = True
            st.rerun()
        else:
            st.error("Contraseña incorrecta.")
    return False


def render_contract_admin(
    data: dict[str, list[dict[str, Any]]], db: SheetClient
) -> None:
    st.subheader("📄 Administración contractual centralizada")
    st.info(
        "Cree, edite o restablezca antecedentes contractuales desde un único lugar. "
        "La eliminación está protegida y queda registrada en el historial."
    )

    establishments = data["establecimientos"]
    contracts = data["contratos"]
    establishment_options = {
        dashboard_name(str(row["nombre"])): row["id"]
        for row in establishments
        if row.get("nombre") and row.get("id") is not None
    }
    if not establishment_options:
        st.error("No existen establecimientos disponibles en Google Sheets.")
        return

    id_to_name = {str(value): key for key, value in establishment_options.items()}
    total_contracts = len(contracts)
    complete_contracts = sum(
        1 for row in contracts
        if row.get("monto_adjudicado") and row.get("fecha_adjudicacion")
        and row.get("duracion_meses")
    )
    validated_contracts = sum(
        1 for row in contracts
        if str(row.get("estado_revision") or "").strip().lower() == "validado"
    )
    metric_a, metric_b, metric_c = st.columns(3)
    metric_a.metric("Antecedentes guardados", total_contracts)
    metric_b.metric("Antecedentes completos", complete_contracts)
    metric_c.metric("Registros validados", validated_contracts)

    filter_a, filter_b = st.columns([1, 2])
    establishment_filter = filter_a.selectbox(
        "Filtrar por establecimiento",
        ["Todos los establecimientos", *establishment_options.keys()],
    )
    search_text = filter_b.text_input(
        "Buscar licitación o instrumento",
        placeholder="Ej.: 1641-121-LR24",
    ).strip().lower()

    filtered_contracts = []
    for contract in contracts:
        contract_establishment = id_to_name.get(
            str(contract.get("establecimiento_id")), "Sin establecimiento"
        )
        if (
            establishment_filter != "Todos los establecimientos"
            and contract_establishment != establishment_filter
        ):
            continue
        if search_text and search_text not in str(contract.get("licitacion") or "").lower():
            continue
        filtered_contracts.append(contract)

    if filtered_contracts:
        preview = [
            {
                "Establecimiento": id_to_name.get(
                    str(row.get("establecimiento_id")), "Sin establecimiento"
                ),
                "Instrumento": row.get("licitacion") or "Sin código",
                "Monto adjudicado": row.get("monto_adjudicado") or "",
                "Fecha adjudicación": row.get("fecha_adjudicacion") or "",
                "Duración (meses)": row.get("duracion_meses") or "",
                "Estado": row.get("estado_revision") or "Borrador",
            }
            for row in filtered_contracts
        ]
        st.dataframe(pd.DataFrame(preview), use_container_width=True, hide_index=True)
    else:
        st.caption("No hay antecedentes guardados que coincidan con los filtros.")

    contract_options: dict[str, dict[str, Any]] = {"➕ Crear nuevo antecedente": {}}
    for contract in filtered_contracts:
        label = (
            f"{contract.get('licitacion', 'Sin código')} · "
            f"{id_to_name.get(str(contract.get('establecimiento_id')), 'Sin establecimiento')}"
        )
        if label in contract_options:
            label = f"{label} · {contract.get('id', '')}"
        contract_options[label] = contract

    selected_label = st.selectbox("Antecedente a gestionar", list(contract_options))
    selected = contract_options[selected_label]
    establishment_names = list(establishment_options)
    default_establishment = (
        establishment_filter
        if establishment_filter in establishment_options
        else establishment_names[0]
    )
    selected_establishment = id_to_name.get(
        str(selected.get("establecimiento_id")), default_establishment
    )

    with st.form("contract_form"):
        establishment = st.selectbox(
            "Establecimiento",
            establishment_names,
            index=establishment_names.index(selected_establishment),
        )
        tender = st.text_input(
            "Licitación / instrumento",
            value=str(selected.get("licitacion") or ""),
        )
        amount = st.number_input(
            "Monto adjudicado",
            min_value=0.0,
            step=100000.0,
            value=float(selected.get("monto_adjudicado") or 0),
            format="%.0f",
        )
        raw_date = selected.get("fecha_adjudicacion")
        try:
            adjudication_date = date.fromisoformat(str(raw_date)) if raw_date else date.today()
        except (TypeError, ValueError):
            adjudication_date = date.today()
        adjudication_date = st.date_input(
            "Fecha de adjudicación", value=adjudication_date
        )
        duration_col, renewal_col = st.columns(2)
        duration = duration_col.number_input(
            "Duración (meses)",
            min_value=0,
            value=int(float(selected.get("duracion_meses") or 0)),
        )
        renewal = renewal_col.number_input(
            "Anticipación de renovación (meses)",
            min_value=0,
            value=int(float(selected.get("anticipacion_meses") or 6)),
        )
        administrative_state = st.selectbox(
            "Estado administrativo",
            ["Vigente", "En renovación", "Finalizado", "Suspendido"],
            index=state_index(
                ["Vigente", "En renovación", "Finalizado", "Suspendido"],
                selected.get("estado_administrativo"),
            ),
        )
        review_state = st.selectbox(
            "Estado del antecedente",
            ["Borrador", "En revisión", "Validado", "Observado"],
            index=state_index(
                ["Borrador", "En revisión", "Validado", "Observado"],
                selected.get("estado_revision"),
            ),
        )
        responsible = st.text_input(
            "Responsable", value=str(selected.get("responsable") or "")
        )
        observations = st.text_area(
            "Observaciones", value=str(selected.get("observaciones") or "")
        )
        submitted = st.form_submit_button(
            "💾 Guardar cambios" if selected.get("id") else "➕ Crear antecedente",
            use_container_width=True,
            type="primary",
        )

    if submitted:
        if not tender.strip():
            st.error("Debe indicar la licitación o instrumento.")
        else:
            contract_id = selected.get("id") or (
                f"ctr-{datetime.now().strftime('%Y%m%d%H%M%S')}-"
                f"{pysecrets.token_hex(3)}"
            )
            payload = {
                "id": contract_id,
                "establecimiento_id": establishment_options[establishment],
                "licitacion": tender.strip(),
                "monto_adjudicado": amount,
                "fecha_adjudicacion": adjudication_date.isoformat(),
                "duracion_meses": duration,
                "anticipacion_meses": renewal,
                "estado_administrativo": administrative_state,
                "responsable": responsible.strip(),
                "observaciones": observations.strip(),
                "actualizado": datetime.now().isoformat(timespec="seconds"),
                "estado_revision": review_state,
                "actualizado_por": "administrador",
            }
            action = "Antecedente actualizado" if selected.get("id") else "Antecedente creado"
            try:
                if selected.get("id"):
                    (
                        db.table("contratos")
                        .update(payload)
                        .eq("id", selected["id"])
                        .execute()
                    )
                else:
                    db.table("contratos").insert(payload).execute()
                db.table("historial_cambios").insert(
                    {
                        "id": (
                            f"hist-{datetime.now().strftime('%Y%m%d%H%M%S')}-"
                            f"{pysecrets.token_hex(3)}"
                        ),
                        "contrato_id": contract_id,
                        "establecimiento_id": establishment_options[establishment],
                        "usuario": "administrador",
                        "accion": action,
                        "detalle": json.dumps(payload, ensure_ascii=False, default=str),
                        "fecha": datetime.now().isoformat(timespec="seconds"),
                    }
                ).execute()
                st.success(f"{action} correctamente.")
                clear_service_cache()
                st.rerun()
            except Exception as exc:
                st.error(f"No fue posible guardar el antecedente: {exc}")

    if selected.get("id"):
        st.divider()
        with st.expander("🛡️ Zona de control: restablecer o eliminar antecedente"):
            instrument = str(selected.get("licitacion") or "")
            st.warning(
                "Esta acción elimina solo los antecedentes complementarios guardados. "
                "El instrumento mensual de origen no se borra y volverá a aparecer como pendiente."
            )
            confirmation = st.text_input(
                f"Para confirmar, escriba exactamente: {instrument}",
                key=f"delete-confirm-{selected.get('id')}",
            )
            delete_clicked = st.button(
                "🗑️ Eliminar antecedente y dejarlo pendiente",
                use_container_width=True,
                disabled=confirmation.strip() != instrument,
            )
            if delete_clicked:
                try:
                    db.table("historial_cambios").insert(
                        {
                            "id": (
                                f"hist-{datetime.now().strftime('%Y%m%d%H%M%S')}-"
                                f"{pysecrets.token_hex(3)}"
                            ),
                            "contrato_id": selected.get("id"),
                            "establecimiento_id": selected.get("establecimiento_id"),
                            "usuario": "administrador",
                            "accion": "Antecedente eliminado",
                            "detalle": json.dumps(
                                selected, ensure_ascii=False, default=str
                            ),
                            "fecha": datetime.now().isoformat(timespec="seconds"),
                        }
                    ).execute()
                    (
                        db.table("contratos")
                        .delete()
                        .eq("id", selected["id"])
                        .execute()
                    )
                    st.success(
                        "Antecedente eliminado. El instrumento quedó nuevamente pendiente."
                    )
                    clear_service_cache()
                    st.rerun()
                except Exception as exc:
                    st.error(f"No fue posible eliminar el antecedente: {exc}")


def find_header_row(uploaded_file, sheet_name: str) -> int:
    preview = pd.read_excel(
        uploaded_file, sheet_name=sheet_name, header=None, nrows=25
    )
    uploaded_file.seek(0)
    for index, row in preview.iterrows():
        cells = {normalize(value) for value in row.tolist() if pd.notna(value)}
        if "establecimiento" in cells and "compromisos" in cells:
            return int(index)
    raise ValueError(
        "No se encontró la fila de encabezados con Establecimiento y Compromisos."
    )


def parse_annex(
    uploaded_file, establishments: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], dict[str, int], list[str], str]:
    excel = pd.ExcelFile(uploaded_file)
    sheet_name = next(
        (
            sheet
            for sheet in excel.sheet_names
            if normalize(sheet) == "anexo n1 minsal"
        ),
        excel.sheet_names[0],
    )
    uploaded_file.seek(0)
    header_row = find_header_row(uploaded_file, sheet_name)
    frame = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header_row)
    uploaded_file.seek(0)
    frame.columns = [str(column).strip() for column in frame.columns]
    columns = {normalize(column): column for column in frame.columns}

    required = [
        "establecimiento",
        "nivel de riesgo",
        "principales causas",
        "medidas implementadas",
        "compromisos",
        "responsable",
        "fecha comprometida",
    ]
    missing = [name for name in required if name not in columns]
    if missing:
        raise ValueError("Faltan columnas: " + ", ".join(missing))

    establishment_ids = {
        dashboard_name(row.get("nombre")): row.get("id")
        for row in establishments
        if row.get("id") is not None
    }
    rows: list[dict[str, Any]] = []
    unmatched: list[str] = []
    counts = {"Rojo": 0, "Amarillo": 0, "Verde": 0}
    detected_period = ""

    def clean(value: Any) -> str:
        return "" if pd.isna(value) else str(value).strip()

    for _, record in frame.iterrows():
        raw_establishment = clean(record.get(columns["establecimiento"]))
        if not raw_establishment:
            continue
        establishment = dashboard_name(raw_establishment)
        establishment_id = establishment_ids.get(establishment)
        if establishment_id is None:
            unmatched.append(raw_establishment)
            continue

        level = level_name(record.get(columns["nivel de riesgo"]))
        if level in counts:
            counts[level] += 1

        if "periodo" in columns and not detected_period:
            detected_period = clean(record.get(columns["periodo"]))

        raw_date = record.get(columns["fecha comprometida"])
        commitment_date = None
        if pd.notna(raw_date):
            try:
                commitment_date = pd.to_datetime(raw_date).date().isoformat()
            except (ValueError, TypeError):
                commitment_date = None

        observations = json.dumps(
            {
                "causas": clean(record.get(columns["principales causas"])),
                "medidas": clean(record.get(columns["medidas implementadas"])),
            },
            ensure_ascii=False,
        )
        row: dict[str, Any] = {
            "establecimiento_id": establishment_id,
            "nivel": level,
            "acciones": clean(record.get(columns["compromisos"])),
            "responsable": clean(record.get(columns["responsable"])),
            "estado": "Publicado",
            "observaciones": observations,
        }
        if commitment_date:
            row["fecha_compromiso"] = commitment_date
        rows.append(row)

    return rows, counts, sorted(set(unmatched)), detected_period


def render_monthly_admin(
    data: dict[str, Any], db: SheetClient
) -> None:
    st.subheader("📦 Actualización mensual del dashboard")
    st.info(
        "Carga el ZIP acumulado de los establecimientos o un CSV de Mercado "
        "Público. La publicación reemplaza la base central vigente para evitar "
        "duplicados y actualiza todos los indicadores del dashboard."
    )

    loads = data.get("cargas_mensuales") or []
    if loads:
        latest = max(loads, key=lambda row: int(row.get("id") or 0))
        st.success(
            f"Base vigente: {latest.get('nombre_archivo') or 'Carga masiva'} · "
            f"{int(latest.get('registros') or 0):,} registros · "
            f"{latest.get('establecimientos') or 0} establecimientos · "
            f"actualizada {latest.get('fecha_carga') or ''}."
        )

    uploaded = st.file_uploader(
        "Seleccionar base masiva", type=["zip", "csv"], key="monthly_database"
    )
    rows: list[dict[str, Any]] = []
    stats: dict[str, Any] = {}

    if uploaded is not None:
        try:
            with st.spinner("Validando y procesando la base..."):
                rows, stats = parse_market_package(
                    uploaded.name, uploaded.getvalue()
                )
            months = stats.get("months") or []
            period = (
                f"{months[0]} a {months[-1]}" if months else "Sin período detectado"
            )
            st.success(
                f"Archivo validado: {stats['files']} CSV · "
                f"{stats['records']:,} registros · "
                f"{len(stats['establishments'])} establecimientos · {period}."
            )
            with st.expander("Ver establecimientos detectados"):
                st.write(", ".join(stats["establishments"]))
        except Exception as exc:
            st.error(f"No fue posible procesar la base: {exc}")

    publish = st.button(
        "☁️ Publicar y actualizar dashboard",
        type="primary",
        use_container_width=True,
        disabled=uploaded is None or not rows,
    )
    if publish and uploaded is not None:
        try:
            with st.spinner("Comprimiendo y guardando la base central..."):
                store_dashboard_dataset(db, rows, uploaded.name, stats)
            st.success(
                f"Base publicada correctamente: {len(rows):,} registros. "
                "El dashboard se actualizará al volver."
            )
            st.cache_data.clear()
            st.rerun()
        except Exception as exc:
            st.error(f"No fue posible publicar la base: {exc}")



REPORT_PERIODS = {
    "Reporte 1": "Enero–Marzo 2026",
    "Reporte 2": "Abril–Junio 2026",
    "Reporte 3": "Julio–Septiembre 2026",
    "Reporte 4": "Octubre–Diciembre 2026",
}


def _official_number(value: Any) -> float:
    text = str(value or "").strip().replace("%", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        try:
            return float(text.replace(".", "").replace(",", "."))
        except ValueError:
            return 0.0


def _official_percentage(value: Any) -> float:
    """Normaliza 61.72, 61,72 o el valor mal interpretado 6172 a 61.72."""
    text = str(value or "").strip().replace("%", "").replace(",", ".")
    try:
        result = float(text)
    except ValueError:
        return 0.0
    while abs(result) > 100 and abs(result) <= 1000000:
        result /= 100
    return round(result, 4)


DEIS_DASHBOARD_NAMES = {
    "110100": "San Juan de Dios",
    "110120": "Félix Bulnes",
    "110130": "Talagante",
    "110150": "Melipilla",
    "110110": "Inst. Traumatológico",
    "110010": "SSMOCC (Dirección)",
    "110300": "CRS S. Allende",
    "110140": "Peñaflor",
    "110160": "Curacaví",
}


@st.cache_data(show_spinner=False)
def parse_minsal_results(file_name, file_bytes, establishments):
    reader = csv.DictReader(
        io.StringIO(_decode_csv(file_bytes)), delimiter=";", quotechar='"'
    )
    if not reader.fieldnames:
        raise ValueError("El CSV no contiene encabezados.")
    headers = {normalize(header): header for header in reader.fieldnames}
    required = ["establecimiento", "denominador", "numerador",
                "trato directo 2026", "nivel de riesgo"]
    missing = [header for header in required if header not in headers]
    if missing:
        raise ValueError("Faltan columnas oficiales: " + ", ".join(missing))
    establishment_ids = {
        dashboard_name(row.get("nombre")): row.get("id")
        for row in establishments if row.get("id") is not None
    }
    output, unmatched = [], []
    for record in reader:
        service_header = headers.get("servicio de salud")
        if service_header and "metropolitano occidente" not in normalize(
                record.get(service_header)):
            continue
        raw_name = str(record.get(headers["establecimiento"]) or "").strip()
        code_deis = str(
            record.get(headers.get("codigo deis", "")) or ""
        ).strip().replace(".0", "")
        canonical_name = DEIS_DASHBOARD_NAMES.get(
            code_deis, dashboard_name(raw_name)
        )
        establishment_id = establishment_ids.get(canonical_name)
        if establishment_id is None:
            unmatched.append(raw_name)
            continue
        output.append({
            "establecimiento_id": establishment_id,
            "codigo_deis": code_deis,
            "denominador": int(round(
                _official_number(record.get(headers["denominador"]))
            )),
            "numerador": int(round(
                _official_number(record.get(headers["numerador"]))
            )),
            # Se guarda con coma como texto RAW para impedir que una planilla
            # con configuración regional chilena convierta 61.72 en 6172.
            "porcentaje_td": (
                f"{_official_percentage(record.get(headers['trato directo 2026'])):.2f}"
                .replace(".", ",")
            ),
            "nivel": level_name(record.get(headers["nivel de riesgo"])),
            "nombre_archivo": file_name,
            "fecha_carga": datetime.now().isoformat(timespec="seconds"),
        })
    if not output:
        raise ValueError("No se encontraron establecimientos del SSMOCC.")
    return output, sorted(set(filter(None, unmatched)))


def render_minsal_admin(data, db):
    st.subheader("⚖️ Resultados oficiales MINSAL")
    st.info(
        "Carga el CSV descargado desde el dashboard MINSAL y asígnalo al "
        "reporte trimestral. La comparación utilizará el mismo período."
    )
    report = st.selectbox("Reporte oficial", list(REPORT_PERIODS),
                          key="minsal_report")
    period = REPORT_PERIODS[report]
    st.text_input("Período asociado", value=period, disabled=True)
    existing = [
        row for row in data.get("resultados_minsal", [])
        if normalize(row.get("reporte")) == normalize(report)
        and normalize(row.get("periodo")) == normalize(period)
    ]
    if existing:
        st.success(
            f"{report} ya tiene {len(existing)} cifras oficiales. "
            "Una nueva publicación reemplazará solo este reporte."
        )
    uploaded = st.file_uploader(
        "Seleccionar CSV oficial MINSAL", type=["csv"],
        key="minsal_results_csv"
    )
    rows, unmatched = [], []
    if uploaded is not None:
        try:
            rows, unmatched = parse_minsal_results(
                uploaded.name, uploaded.getvalue(), data["establecimientos"])
            names = {row.get("id"): dashboard_name(row.get("nombre"))
                     for row in data["establecimientos"]}
            st.success(f"Archivo validado: {len(rows)} establecimientos.")
            st.dataframe([{
                "Establecimiento": names.get(row["establecimiento_id"], ""),
                "% TD MINSAL": row["porcentaje_td"],
                "Nivel": row["nivel"],
                "Numerador": row["numerador"],
                "Denominador": row["denominador"],
            } for row in rows], use_container_width=True, hide_index=True)
            if unmatched:
                st.warning("No se reconocieron: " + ", ".join(unmatched))
        except Exception as exc:
            st.error(f"No fue posible procesar el archivo MINSAL: {exc}")
    publish = st.button(
        "⚖️ Publicar resultados MINSAL en Google Sheets",
        type="primary", use_container_width=True,
        disabled=uploaded is None or not rows
    )
    if publish:
        try:
            # Conserva los demás reportes y reemplaza el seleccionado mediante
            # una única escritura. Evita leer/borrar fila por fila y agotar la
            # cuota de Google Sheets.
            retained = [
                dict(current)
                for current in data.get("resultados_minsal", [])
                if not (
                    normalize(current.get("reporte")) == normalize(report)
                    and normalize(current.get("periodo")) == normalize(period)
                )
            ]
            published_rows = []
            next_id = max(
                [int(row.get("id") or 0) for row in retained]
                + [0]
            ) + 1
            for row in rows:
                published = dict(row)
                published["id"] = next_id
                published["reporte"] = report
                published["periodo"] = period
                published_rows.append(published)
                next_id += 1
            db.replace_records(
                "resultados_minsal", retained + published_rows
            )
            st.success(
                f"{report} publicado con {len(published_rows)} cifras MINSAL. "
                "Todos los establecimientos quedaron homologados."
            )
            st.cache_data.clear()
            st.rerun()
        except Exception as exc:
            st.error(f"No fue posible publicar los resultados MINSAL: {exc}")


def render_plan_admin(
    data: dict[str, list[dict[str, Any]]], db: SheetClient
) -> None:
    st.subheader("☁️ Plan de trabajo oficial · Anexo N°1")
    st.info(
        "La clasificación del Anexo es oficial y se muestra en el Plan de trabajo. "
        "El semáforo principal mantiene el cálculo automático por % de Trato Directo."
    )

    current = latest_plan(data["planes"])
    if current:
        st.success(
            f"Plan vigente: {current.get('nombre_archivo') or 'Anexo N°1'} · "
            f"{current.get('establecimientos') or 0} establecimientos · "
            f"{current.get('rojos') or 0} rojos · "
            f"{current.get('amarillos') or 0} amarillos · "
            f"{current.get('verdes') or 0} verdes."
        )

    uploaded = st.file_uploader(
        "Seleccionar Anexo N°1", type=["xlsx", "xls"]
    )
    col1, col2 = st.columns(2)
    report = col1.text_input("Reporte", value="Reporte 1")
    period = col2.text_input("Período", value="Enero–Marzo 2026")
    publication_date = st.date_input("Fecha de publicación", value=date.today())

    rows: list[dict[str, Any]] = []
    counts = {"Rojo": 0, "Amarillo": 0, "Verde": 0}
    unmatched: list[str] = []
    detected_period = ""

    if uploaded is not None:
        try:
            rows, counts, unmatched, detected_period = parse_annex(
                uploaded, data["establecimientos"]
            )
            st.success(
                f"Archivo validado: {len(rows)} establecimientos · "
                f"{counts['Rojo']} rojos · {counts['Amarillo']} amarillos · "
                f"{counts['Verde']} verdes."
            )
            if unmatched:
                st.warning("No se reconocieron: " + ", ".join(unmatched))
        except Exception as exc:
            st.error(f"No fue posible procesar el Anexo: {exc}")

    publish = st.button(
        "☁️ Publicar plan oficial en Google Sheets",
        type="primary",
        use_container_width=True,
        disabled=uploaded is None or not rows,
    )

    if publish and uploaded is not None:
        metadata = {
            "nombre_archivo": uploaded.name,
            "reporte": report.strip(),
            "periodo": (detected_period or period).strip(),
            "fecha_publicacion": publication_date.isoformat(),
            "establecimientos": len(rows),
            "rojos": counts["Rojo"],
            "amarillos": counts["Amarillo"],
            "verdes": counts["Verde"],
            "url_archivo": "",
        }
        try:
            # Etiqueta las filas antiguas del primer reporte para no perderlas.
            existing_plans = db.table("planes").select("*").execute().data or []
            previous = latest_plan(existing_plans)
            existing_rows = db.table("plan_trabajo").select("*").execute().data or []
            if previous:
                for record in existing_rows:
                    if (
                        record.get("id") is not None
                        and not record.get("reporte")
                        and not record.get("periodo")
                    ):
                        db.table("plan_trabajo").update({
                            "reporte": str(previous.get("reporte") or ""),
                            "periodo": str(previous.get("periodo") or ""),
                            "fecha_publicacion": str(
                                previous.get("fecha_publicacion") or ""
                            ),
                        }).eq("id", record["id"]).execute()
                existing_rows = (
                    db.table("plan_trabajo").select("*").execute().data or []
                )

            # Republicar reemplaza únicamente el mismo reporte y período.
            for record in existing_rows:
                if (
                    record.get("id") is not None
                    and normalize(record.get("reporte")) == normalize(metadata["reporte"])
                    and normalize(record.get("periodo")) == normalize(metadata["periodo"])
                ):
                    db.table("plan_trabajo").delete().eq(
                        "id", record["id"]
                    ).execute()
            for record in existing_plans:
                if (
                    record.get("id") is not None
                    and normalize(record.get("reporte")) == normalize(metadata["reporte"])
                    and normalize(record.get("periodo")) == normalize(metadata["periodo"])
                ):
                    db.table("planes").delete().eq("id", record["id"]).execute()

            for row in rows:
                row["reporte"] = metadata["reporte"]
                row["periodo"] = metadata["periodo"]
                row["fecha_publicacion"] = metadata["fecha_publicacion"]

            db.table("plan_trabajo").insert(rows).execute()
            db.table("planes").insert(metadata).execute()
            st.success(
                f"Plan publicado: {len(rows)} establecimientos guardados en Google Sheets."
            )
            st.rerun()
        except Exception as exc:
            st.error(f"No fue posible publicar el plan: {exc}")

    st.caption(
        f"Registros actualmente disponibles en plan_trabajo: {len(data['plan_trabajo'])}."
    )


def render_admin(data: dict[str, list[dict[str, Any]]]) -> None:
    if not admin_login():
        return
    db = service_client()
    if db is None:
        st.error("Falta configurar gcp_service_account / GSHEET_ID en Streamlit Secrets.")
        return

    left, right = st.columns([6, 1])
    left.success("Sesión administrativa activa.")
    if right.button("Cerrar sesión"):
        st.session_state["admin_authenticated"] = False
        st.query_params.clear()
        st.rerun()

    contracts_tab, bulk_tab, monthly_tab, minsal_tab, plan_tab = st.tabs(
        ["📄 Gestión contractual", "📥 Carga masiva",
         "📦 Actualización mensual", "⚖️ Resultados MINSAL", "☁️ Plan oficial"]
    )
    with contracts_tab:
        render_contract_admin(data, db)
    with bulk_tab:
        render_contract_bulk_admin(data, db)
    with monthly_tab:
        render_monthly_admin(data, db)
    with minsal_tab:
        render_minsal_admin(data, db)
    with plan_tab:
        render_plan_admin(data, db)

    if st.button("← Volver al dashboard"):
        st.query_params.clear()
        st.rerun()


# -----------------------------------------------------------------------------
# HTML: INTEGRACIÓN NATIVA, SIN OBSERVADORES NI TEMPORIZADORES
# -----------------------------------------------------------------------------
def load_html() -> str:
    for path in HTML_FILES:
        if path.exists():
            return path.read_text(encoding="utf-8")
    raise FileNotFoundError("No se encontró index.html junto a app.py.")


def replace_native_loader(html: str, function_name: str, replacement: str) -> str:
    if function_name == "loadPlan":
        pattern = re.compile(
            r"function\s+loadPlan\(\)\s*\{\s*"
            r"try\s*\{\s*PLAN\s*=\s*JSON\.parse\(localStorage\.getItem\(PLAN_KEY\)\)"
            r"\s*\|\|\s*\{meta:null,items:\[\]\};\s*\}\s*"
            r"catch\(e\)\s*\{\s*PLAN\s*=\s*\{meta:null,items:\[\]\};\s*\}\s*\}",
            flags=re.DOTALL,
        )
    else:
        pattern = re.compile(
            r"function\s+loadLic\(\)\s*\{\s*"
            r"try\s*\{\s*LIC\s*=\s*JSON\.parse\(localStorage\.getItem\(LIC_KEY\)\s*\|\|\s*'\{\}'\);\s*\}"
            r"\s*catch\(e\)\s*\{\s*LIC\s*=\s*\{\};\s*\}\s*\}",
            flags=re.DOTALL,
        )
    return pattern.sub(replacement, html, count=1)


def replace_admin_button(html: str) -> str:
    pattern = re.compile(
        r"<button(?P<attrs>[^>]*)>(?P<body>(?:(?!</button>).)*?Administrador(?:(?!</button>).)*?)</button>",
        flags=re.IGNORECASE | re.DOTALL,
    )

    def replacement(match: re.Match[str]) -> str:
        attrs = re.sub(
            r"\s+onclick\s*=\s*(['\"]).*?\1",
            "",
            match.group("attrs"),
            flags=re.IGNORECASE | re.DOTALL,
        )
        return (
            f'<a{attrs} href="https://td-ssmocc.streamlit.app/?admin=1" '
            f'target="_blank" rel="noopener noreferrer">'
            f'{match.group("body")}</a>'
        )

    return pattern.sub(replacement, html, count=1)


def inject_native_data(
    html: str,
    contract_payload: dict[str, dict[str, Any]],
    plan_payload: dict[str, Any],
    plan_history: list[dict[str, Any]],
    dataset_gzip_b64: str,
    coverage_period_max: str,
) -> str:
    contracts_json = json.dumps(
        contract_payload, ensure_ascii=False, separators=(",", ":")
    ).replace("</", "<\\/")
    plan_json = json.dumps(
        plan_payload, ensure_ascii=False, separators=(",", ":")
    ).replace("</", "<\\/")
    history_json = json.dumps(
        plan_history, ensure_ascii=False, separators=(",", ":")
    ).replace("</", "<\\/")

    preload = f"""
    <script>
      window.__SHEETS_CONTRACTS__ = {contracts_json};
      window.__SHEETS_PLAN__ = {plan_json};
      window.__SHEETS_PLAN_HISTORY__ = {history_json};
      window.__SHEETS_ROWS_GZIP__ = {json.dumps(dataset_gzip_b64)};
      window.__SHEETS_COVERAGE_MAX__ = {json.dumps(coverage_period_max)};

      function ssmoccPublishedReportNumbers() {{
        return new Set((window.__SHEETS_PLAN_HISTORY__ || []).map(plan => {{
          const label = String((plan.meta || {{}}).reporte || '');
          const match = label.match(/(\\d+)/);
          return match ? Number(match[1]) : null;
        }}).filter(Boolean));
      }}

      function ssmoccRefreshReportStatus() {{
        const calendar = document.getElementById('td-calendar');
        if (!calendar) return;
        const published = ssmoccPublishedReportNumbers();
        Array.from(calendar.children).forEach((card, index) => {{
          const reportNumber = index + 1;
          if (!published.has(reportNumber) || card.dataset.reportPublished === '1') {{
            return;
          }}
          card.dataset.reportPublished = '1';
          card.style.borderLeftColor = '#2e7d32';
          const header = card.children[0];
          const badge = header && header.children[1];
          if (badge) {{
            badge.style.background = '#2e7d321a';
            badge.style.color = '#2e7d32';
            badge.innerHTML = '<i class="fa-solid fa-circle-check mr-1"></i>Entregado';
          }}
          const statusLine = card.children[3];
          if (statusLine) {{
            statusLine.style.color = '#2e7d32';
            statusLine.textContent = 'Reporte cerrado y antecedentes remitidos';
          }}
        }});
      }}

      function ssmoccInstallReconTabs() {{
        const tbody = document.getElementById('td-recon');
        if (!tbody) return;
        const card = tbody.closest('.bg-white');
        if (!card || card.querySelector('#td-recon-tabs')) return;
        const intro = card.querySelector('p');
        const tabs = document.createElement('div');
        tabs.id = 'td-recon-tabs';
        tabs.className = 'ssmocc-recon-tabs';
        const periods = ['Enero–Marzo', 'Abril–Junio',
                         'Julio–Septiembre', 'Octubre–Diciembre'];
        const published = ssmoccPublishedReportNumbers();
        tabs.innerHTML = periods.map((period, index) => {{
          const number = index + 1;
          const available = published.has(number);
          return '<button type="button" data-recon-report="' + number +
            '" class="ssmocc-recon-tab">' +
            '<span class="ssmocc-recon-number">' + number + '° Reporte</span>' +
            '<span class="ssmocc-recon-period">' + period + '</span>' +
            '<span class="ssmocc-recon-status ' +
            (available ? 'available' : 'pending') + '">' +
            (available ? 'Resultados disponibles' : 'Pendiente') +
            '</span></button>';
        }}).join('');
        if (intro) intro.insertAdjacentElement('afterend', tabs);
        tabs.querySelectorAll('[data-recon-report]').forEach(button => {{
          button.addEventListener('click', function() {{
            const reportNumber = Number(button.dataset.reconReport);
            const cards = document.querySelectorAll('#td-calendar > div');
            if (cards[reportNumber - 1]) cards[reportNumber - 1].click();
            ssmoccRefreshReconTabs(reportNumber);
          }});
        }});
        ssmoccRefreshReconTabs(window.__SHEETS_ACTIVE_REPORT__ || 1);
      }}

      function ssmoccRefreshReconTabs(activeReport) {{
        document.querySelectorAll('[data-recon-report]').forEach(button => {{
          button.classList.toggle(
            'active', Number(button.dataset.reconReport) === Number(activeReport)
          );
        }});
      }}

      document.addEventListener('DOMContentLoaded', function() {{
        const observer = new MutationObserver(function() {{
          ssmoccRefreshReportStatus();
          ssmoccInstallReconTabs();
        }});
        observer.observe(document.body, {{childList:true, subtree:true}});
        ssmoccRefreshReportStatus();
        ssmoccInstallReconTabs();
      }});

      document.addEventListener('click', function(event) {{
        const card = event.target.closest('#td-calendar > div');
        if (!card) return;
        const cards = Array.from(card.parentElement.children);
        const reportNumber = cards.indexOf(card) + 1;
        const history = window.__SHEETS_PLAN_HISTORY__ || [];
        const selected = history.find(plan => {{
          const label = String((plan.meta || {{}}).reporte || '');
          const match = label.match(/(\\d+)/);
          return match && Number(match[1]) === reportNumber;
        }});
        if (!selected) {{
          window.__SHEETS_ACTIVE_REPORT__ = reportNumber;
          const periods = {{
            1: 'Enero–Marzo', 2: 'Abril–Junio',
            3: 'Julio–Septiembre', 4: 'Octubre–Diciembre'
          }};
          window.__SHEETS_PLAN__ = {{
            meta: {{
              reporte: 'Reporte ' + reportNumber,
              periodo: periods[reportNumber] + ' 2026'
            }},
            items: []
          }};
          if (typeof PLAN !== 'undefined') PLAN = window.__SHEETS_PLAN__;
          if (typeof tdState !== 'undefined') {{
            tdState.period = 'q' + reportNumber;
            tdState.planSel = null;
          }}
          if (typeof renderTD === 'function') renderTD();
          ssmoccRefreshReconTabs(reportNumber);
          const selector = document.getElementById('td-plan-sel');
          if (selector) {{
            selector.innerHTML =
              '<option>Reporte ' + reportNumber + ' · Plan no disponible</option>';
            selector.disabled = true;
          }}
          const planBox = document.getElementById('td-plan');
          if (planBox) {{
            planBox.innerHTML =
              '<div class="border border-slate-200 bg-slate-50 rounded-xl px-5 py-6 text-center">' +
              '<div class="w-12 h-12 mx-auto mb-3 rounded-full bg-blue-50 text-govblue grid place-items-center">' +
              '<i class="fa-solid fa-calendar-clock text-xl"></i></div>' +
              '<div class="font-tight font-bold text-slate-800 text-[17px]">Plan aún no disponible</div>' +
              '<p class="text-[13px] text-slate-500 mt-2 max-w-xl mx-auto">' +
              'El Anexo N°1 correspondiente al ' + reportNumber + '° Reporte (' +
              periods[reportNumber] + ') todavía no ha sido publicado. ' +
              'La información estará disponible una vez finalizado el proceso de recepción y validación.</p>' +
              '</div>';
          }}
          Array.from(card.parentElement.children).forEach((item, index) => {{
            item.classList.toggle(
              'ssmocc-report-selected', index + 1 === reportNumber
            );
          }});
          if (typeof toast === 'function') {{
            toast('El plan del reporte ' + reportNumber + ' aún no está disponible');
          }}
          return;
        }}
        const selector = document.getElementById('td-plan-sel');
        if (selector) selector.disabled = false;
        window.__SHEETS_PLAN__ = selected;
        window.__SHEETS_ACTIVE_REPORT__ = reportNumber;
        ssmoccRefreshReconTabs(reportNumber);
        if (typeof PLAN !== 'undefined') PLAN = selected;
        if (typeof tdState !== 'undefined') {{
          tdState.planSel = null;
          tdState.period = 'q' + reportNumber;
        }}
        if (typeof renderTD === 'function') renderTD();
        setTimeout(function() {{
          const currentCards = document.querySelectorAll('#td-calendar > div');
          currentCards.forEach((item, index) => {{
            item.classList.toggle(
              'ssmocc-report-selected', index + 1 === reportNumber
            );
          }});
        }}, 0);
      }});
    </script>
    """
    html = html.replace("</head>", preload + "\n</head>", 1)

    html = html.replace(
        "META=D.meta.map(m=>Object.assign({},m)); BASE=D.rows;",
        "META=D.meta.map(m=>Object.assign({},m));"
        "if(window.__SHEETS_ROWS_GZIP__){"
        "const centralBin=Uint8Array.from(atob(window.__SHEETS_ROWS_GZIP__),"
        "c=>c.charCodeAt(0));"
        "const centralStream=new Blob([centralBin]).stream()"
        ".pipeThrough(new DecompressionStream('gzip'));"
        "BASE=JSON.parse(await new Response(centralStream).text());"
        "const coverage=String(window.__SHEETS_COVERAGE_MAX__||'');"
        "if(/^\\d{4}-\\d{2}$/.test(coverage)){"
        "BASE=BASE.filter(r=>!r.f||String(r.f).slice(0,7)<=coverage);"
        "}"
        "}else{BASE=D.rows;}",
        1,
    )

    html = html.replace(
        "await loadExtra();",
        "if(window.__SHEETS_ROWS_GZIP__){"
        "EXTRA=[];BATCHES=[];"
        "try{localStorage.removeItem(ROWS_KEY);"
        "localStorage.removeItem(BATCH_KEY);}catch(e){}"
        "}else{await loadExtra();}",
        1,
    )

    html = replace_native_loader(
        html,
        "loadLic",
        "function loadLic(){try{LIC=window.__SHEETS_CONTRACTS__||{};}catch(e){LIC={};}}",
    )
    html = replace_native_loader(
        html,
        "loadPlan",
        "function loadPlan(){try{PLAN=window.__SHEETS_PLAN__||{meta:null,items:[]};}catch(e){PLAN={meta:null,items:[]};}}",
    )

    # El semáforo conserva el porcentaje calculado, pero utiliza la clasificación
    # oficial del Anexo N°1 cuando existe un plan publicado en Google Sheets.
    automatic_level = (
        "rows.push({e:k,total:a.total,td:a.td,nTD:a.nTD,pct,"
        "nivel:tdNivel(pct)}); });"
    )
    official_level = (
        "const official=(PLAN.items||[]).find(it=>matchEstab(it.estab)===k);"
        "const officialNivel=official?normNivel(official.nivel):'';"
        "const nivel=['Rojo','Amarillo','Verde'].includes(officialNivel)"
        "?officialNivel:tdNivel(pct);"
        "rows.push({e:k,total:a.total,td:a.td,nTD:a.nTD,pct,nivel}); });"
    )
    html = html.replace(automatic_level, official_level, 1)

    # Respaldo: inicializa PLAN desde Google Sheets incluso si cambia el formato del loader.
    html = re.sub(
        r"let\s+PLAN\s*=\s*\{meta:null,items:\[\]\};",
        "let PLAN=window.__SHEETS_PLAN__||{meta:null,items:[]};",
        html,
        count=1,
    )

    html = replace_admin_button(html)
    html = html.replace(
        "document.getElementById('admin-btn').onclick=openAdmin;",
        "document.getElementById('admin-btn').onclick=()=>{"
        "window.top.location.href='https://td-ssmocc.streamlit.app/?admin=1';"
        "};",
        1,
    )
    html = html.replace(
        "Ordenado por % TD. Seleccione una fila para ver el detalle.",
        "Ordenado por % TD. Se utiliza el nivel oficial del Anexo N°1 cuando está publicado; de lo contrario, se calcula según los umbrales.",
        1,
    )
    return html


def apply_layout_patch(html: str) -> str:
    patch = """
    <style>
      html, body {
        width: 100% !important;
        max-width: 100% !important;
        overflow-x: hidden !important;
      }
      [class~="max-w-[1600px]"] { max-width: 100% !important; }
      #td-calendar > div {
        cursor: pointer !important;
        position: relative;
        transition: transform .15s ease, box-shadow .15s ease;
      }
      #td-calendar > div:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 18px rgba(15, 23, 42, .10);
      }
      #td-calendar > div::after {
        content: "Seleccionar reporte";
        display: block;
        margin-top: 8px;
        font-size: 11px;
        font-weight: 700;
        color: #0063af;
      }
      #td-calendar > div.ssmocc-report-selected {
        outline: 3px solid rgba(0, 99, 175, .28);
        background: #f0f8ff;
      }
      #td-calendar > div.ssmocc-report-selected::after {
        content: "Reporte seleccionado";
      }
      .ssmocc-recon-tabs {
        display: grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: 10px;
        margin: 16px 0 8px;
      }
      .ssmocc-recon-tab {
        text-align: left;
        border: 1px solid #e2e8f0;
        border-radius: 12px;
        padding: 11px 12px;
        background: #f8fafc;
        transition: border-color .15s, box-shadow .15s, background .15s;
      }
      .ssmocc-recon-tab:hover {
        border-color: #7fb3e0;
        background: #f0f8ff;
      }
      .ssmocc-recon-tab.active {
        border-color: #0063af;
        background: #eef7ff;
        box-shadow: 0 0 0 2px rgba(0,99,175,.14);
      }
      .ssmocc-recon-number {
        display: block;
        color: #00305e;
        font-size: 13px;
        font-weight: 800;
      }
      .ssmocc-recon-period {
        display: block;
        color: #64748b;
        font-size: 11px;
        margin-top: 2px;
      }
      .ssmocc-recon-status {
        display: inline-block;
        margin-top: 7px;
        padding: 2px 7px;
        border-radius: 999px;
        font-size: 10px;
        font-weight: 700;
      }
      .ssmocc-recon-status.available {
        color: #2e7d32;
        background: rgba(46,125,50,.10);
      }
      .ssmocc-recon-status.pending {
        color: #64748b;
        background: #e2e8f0;
      }
      @media (max-width: 700px) {
        .ssmocc-recon-tabs { grid-template-columns: repeat(2, 1fr); }
      }
      @media (min-width: 1024px) {
        #sidebar {
          position: static !important;
          inset: auto !important;
          transform: none !important;
          max-height: none !important;
          height: auto !important;
          overflow: visible !important;
          align-self: flex-start !important;
        }
      }
    </style>
    """
    return html.replace("</head>", patch + "\n</head>", 1)


def main() -> None:
    data = load_data()
    admin_requested = str(st.query_params.get("admin", "0")) == "1"
    portal_requested = str(st.query_params.get("portal", "0")) == "1"

    if admin_requested or st.session_state.get("admin_authenticated"):
        render_admin(data)
        st.stop()
    if portal_requested or st.session_state.get("portal_user_id"):
        db = service_client()
        if db is None:
            st.error("No existe conexión con Google Sheets.")
        else:
            render_establishment_portal(data, db)
        st.stop()

    try:
        html = load_html()
    except FileNotFoundError as exc:
        st.error(str(exc))
        st.stop()

    html = inject_native_data(
        html,
        contracts_for_html(data["contratos"], data["establecimientos"]),
        attach_minsal_results(
            plan_for_html(
                data["planes"], data["plan_trabajo"], data["establecimientos"]
            ),
            data["resultados_minsal"], data["establecimientos"],
        ),
        plan_history_for_html(
            data["planes"], data["plan_trabajo"], data["establecimientos"],
            data["resultados_minsal"],
        ),
        str(data.get("dataset_gzip_b64") or ""),
        (
            date.today().replace(day=1) - timedelta(days=1)
        ).strftime("%Y-%m"),
    )
    html = apply_layout_patch(html)
    components.html(html, height=4300, scrolling=False)


if __name__ == "__main__":
    main()
